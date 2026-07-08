"""
Mandanten-Pipeline: Hankook Tire Austria GmbH.

Besonderheit gegenüber der Standard-Pipeline:
  * Der Anhang weist die sonstigen finanziellen Verpflichtungen getrennt aus:
        Verpflichtungen aus Leasingverträgen   (= KFZ / "Car lease")
        Verpflichtungen aus Mietverträgen      (= "Leasing Warehouse" + "Rent Office")
    jeweils Spalte 1 = folgendes Geschäftsjahr, Spalte 2 = folgende fünf Jahre.
  * Der Nachweis liegt als Excel ("Eventualverbindlichkeiten … Miete, Fuhrpark")
    mit dem Summenblatt "Zusammenfassung Hankook" vor:
        Leasing Warehouse | Rent Office | Car lease | Total

Reconciliation (Spalte "folgendes Geschäftsjahr"):
    Anhang "Verpflichtungen aus Leasingverträgen"  ==  Excel "Car lease"
    Anhang "Verpflichtungen aus Mietverträgen"     ==  Excel "Leasing Warehouse" + "Rent Office"

Das "Hirn" (pruefung.comparator) summiert die Belegfakten je Typ und vergleicht
gegen den Anhangwert – daher liefern wir für Miete ZWEI Fakten desselben Typs.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Optional

import openpyxl
import pdfplumber

from ..pruefung.extractor import AnhangPosition, ExtractedFact, _trailing_numbers, _parse_de_number
from ..vorjahresvergleich.extractor import AnhangItem
from .base import Pipeline

# Belegtypen dieser Pipeline
TYPE_LEASING = "hankook_leasing"   # KFZ-Leasing ("Car lease")
TYPE_MIETE = "hankook_miete"       # Miete/Warehouse ("Leasing Warehouse" + "Rent Office")
TYPE_EVENTUALV = "hankook_eventualverbindlichkeiten"  # das Summen-Excel

# Latente Steuern: Detail-PDF + je Komponente ein eigener Typ (komponentenweiser Abgleich)
TYPE_LST_DETAIL = "hankook_latente_steuern"   # das Detail-PDF "lat. Steuern"
TYPE_LST_LEASING = "hankook_lst_leasing_kfz"
TYPE_LST_FIRMENWERT = "hankook_lst_firmenwert"
TYPE_LST_JUBILAEUM = "hankook_lst_jubilaeum"

# Deutsches Zahlenformat (für die Latente-Steuern-Parser)
_NUM_RE = re.compile(r"\(?-?\d{1,3}(?:\.\d{3})*(?:,\d+)?\)?")

# Komponenten der latenten Steuern: (Anker im Text, Section, Belegtyp, Label)
_LST_COMPONENTS = [
    (r"Leasing\s*KFZ", "LatSteuer_Leasing_KFZ", TYPE_LST_LEASING, "Latente Steuer Aktivposten Leasing KFZ"),
    (r"Firmen",        "LatSteuer_Firmenwert",  TYPE_LST_FIRMENWERT, "Latente Steuer Geschäfts-(Firmen-)wert"),
    (r"Jubil",         "LatSteuer_Jubilaeum",   TYPE_LST_JUBILAEUM, "Latente Steuer Rückstellung Jubiläumsgelder"),
]

# Vorjahresvergleich (Modus 1): Vorwärts-Verpflichtungen sind KEINE Kontinuitätsposten
_VJV_EXCLUDE_RE = re.compile(r"verpflichtungen\s+aus\s+(leasing|miet)", re.I)


def _is_hankook_vjv_item(it: AnhangItem) -> bool:
    """True, wenn der Posten in den Vorjahresvergleich (Bilanzkontinuität) gehört.

    Beseitigt die Fehlerquellen des Standard-Extraktors bei Hankook:
      * Nutzungsdauer-Angaben (z.B. Firmenwert "10", BGA "5" = JAHRE) sowie
        Seitenzahlen/Prosa-Zahlen erscheinen als Einzelwert und kollidieren
        mit gleichlautenden Anlagenspiegel-Labels -> Fehlmatch.
      * Verpflichtungen aus Leasing-/Mietverträgen sind Vorwärtsangaben
        (folgendes Jahr / folgende 5 Jahre), keine Bilanzkontinuität; sie
        werden in Modus 2 geprüft.
    """
    if _VJV_EXCLUDE_RE.search(it.label):
        return False
    if it.double_row or it.prior_values:   # Anlagenspiegel / Posten mit Vorjahreswert
        return True
    return len(it.current_values) >= 3     # einzeiliger Spiegel ja, Einzel-Rauschzahl nein


# ---------------------------------------------------------------------------
# Anhang-Seite: die beiden Verpflichtungs-Zeilen aus dem Hankook-Anhang lesen
# ---------------------------------------------------------------------------
def _iter_pdf_pages(pdf_path: Path):
    """Liefert (Seitenzahl, Seitentext). Robust: bei Lesefehlern leere Folge."""
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                try:
                    yield page_num, (page.extract_text() or "")
                except Exception:
                    yield page_num, ""
    except Exception:
        return


def _extract_hankook_anhang_positions(pdf_path: Path) -> list[AnhangPosition]:
    positions: list[AnhangPosition] = []
    for page_num, text in _iter_pdf_pages(pdf_path):
        for line in text.split("\n"):
            if re.search(r"Verpflichtungen\s+aus\s+Leasingvertr", line, re.I):
                nums = _trailing_numbers(line)
                if nums:
                    positions.append(AnhangPosition(
                        section="Verpflichtungen_Leasing",
                        label="Verpflichtungen aus Leasingverträgen (folgendes GJ)",
                        current_value=nums[0],
                        prior_value=nums[1] if len(nums) > 1 else None,
                        page=page_num,
                        source_line=line.strip(),
                    ))
            elif re.search(r"Verpflichtungen\s+aus\s+Mietvertr", line, re.I):
                nums = _trailing_numbers(line)
                if nums:
                    positions.append(AnhangPosition(
                        section="Verpflichtungen_Miete",
                        label="Verpflichtungen aus Mietverträgen (folgendes GJ)",
                        current_value=nums[0],
                        prior_value=nums[1] if len(nums) > 1 else None,
                        page=page_num,
                        source_line=line.strip(),
                    ))
    return positions


# ---------------------------------------------------------------------------
# Beleg-Seite: das Summenblatt "Zusammenfassung" auslesen
# ---------------------------------------------------------------------------
def _first_number(row) -> Optional[float]:
    """Erster echter Zahlenwert einer Zeile (= Spalte 'folgendes Geschäftsjahr')."""
    for v in row:
        if isinstance(v, bool):
            continue
        if isinstance(v, (int, float)):
            return float(v)
    return None


def _categorize(label_lower: str) -> Optional[str]:
    if "car" in label_lower and "lease" in label_lower:
        return TYPE_LEASING
    if "warehouse" in label_lower:
        return TYPE_MIETE
    if "office" in label_lower or "rent" in label_lower or "miet" in label_lower:
        return TYPE_MIETE
    return None


def _extract_hankook_eventualverbindlichkeiten(xlsx_path: Path) -> list[ExtractedFact]:
    facts: list[ExtractedFact] = []
    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    except Exception:
        return facts

    # Summenblatt bevorzugen
    sheets = sorted(
        wb.worksheets,
        key=lambda ws: 0 if "zusammenfassung" in ws.title.lower() else 1,
    )
    for ws in sheets:
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            label = next(
                (str(v).strip() for v in cells if isinstance(v, str) and v.strip()),
                "",
            )
            if not label:
                continue
            low = label.lower()
            if low.startswith(("total", "summe", "gesamt")):
                continue  # Summenzeile überspringen – wir prüfen die Komponenten
            ctype = _categorize(low)
            if ctype is None:
                continue
            value = _first_number(cells)
            if value is None:
                continue
            facts.append(ExtractedFact(
                source_type=ctype,
                label=label,
                value=value,
                filename=Path(xlsx_path).name,
                sheet=ws.title,
            ))
        if facts:
            break  # Summenblatt gefunden und ausgewertet
    wb.close()
    return facts


# ---------------------------------------------------------------------------
# Latente Steuern – Anhang (Fließtext Seite 4) und Detail-PDF
# ---------------------------------------------------------------------------
def _parse_lst_anhang(text: str, page_num: int = 1) -> list[AnhangPosition]:
    """Reine Parselogik (testbar): drei Steuerlatenz-Beträge aus dem Fließtext.

    Muster: '… Steuerlatenzen für den Aktivposten Leasing KFZ in Höhe von
    EUR 13.198,00, des Geschäfts-(Firmen-)wertes … EUR 137.097,00 und der
    Rückstellung für Jubiläumsgelder … EUR 1.846,00.'
    """
    positions: list[AnhangPosition] = []
    i = text.lower().find("steuerlatenz")
    if i < 0:
        return positions
    window = text[i:i + 400]  # nur der eine Satz, nicht die Rückstellungstabelle
    for anchor, section, _stype, label in _LST_COMPONENTS:
        m = re.search(anchor + r".*?EUR\s*(" + _NUM_RE.pattern + r")", window, re.I)
        if not m:
            continue
        val = _parse_de_number(m.group(1))
        if val is not None:
            positions.append(AnhangPosition(
                section=section, label=label,
                current_value=val, prior_value=None,
                page=page_num, source_line=m.group(0).strip()[:120],
            ))
    return positions


def _parse_lst_beleg(text: str, filename: str = "") -> list[ExtractedFact]:
    """Reine Parselogik (testbar): je Komponente die latente Steuer (4. Spalte).

    Zeilenform: '<Bezeichnung> <UGB> <StR> <Differenz> <latente Steuer>'.
    """
    facts: list[ExtractedFact] = []
    for anchor, _section, stype, label in _LST_COMPONENTS:
        m = re.search(anchor, text, re.I)
        if not m:
            continue
        after = text[m.end():m.end() + 120]
        nums = [v for x in _NUM_RE.finditer(after) if (v := _parse_de_number(x.group())) is not None]
        if not nums:
            continue
        value = nums[3] if len(nums) >= 4 else nums[-1]  # 4. Spalte = latente Steuer
        facts.append(ExtractedFact(source_type=stype, label=label, value=value, filename=filename))
    return facts


def _pdf_text(pdf_path: Path) -> str:
    """Gesamten PDF-Text zeilenweise zu einer Zeile je Seite zusammenführen (robust)."""
    return " ".join(
        " ".join(text.split("\n")) for _page, text in _iter_pdf_pages(pdf_path)
    )


def _extract_hankook_latente_steuern_anhang(pdf_path: Path) -> list[AnhangPosition]:
    for page_num, raw in _iter_pdf_pages(pdf_path):
        text = " ".join(raw.split("\n"))
        if "steuerlatenz" in text.lower():
            return _parse_lst_anhang(text, page_num)
    return []


def _extract_hankook_latente_steuern_beleg(pdf_path: Path) -> list[ExtractedFact]:
    return _parse_lst_beleg(_pdf_text(pdf_path), Path(pdf_path).name)


# ---------------------------------------------------------------------------
# Belegerkennung dieser Pipeline
# ---------------------------------------------------------------------------
def _detect_hankook_type(path: Path) -> str:
    """Erkennt die Hankook-Belege: Eventualverbindlichkeiten-Excel und lat.-Steuern-PDF."""
    ext = Path(path).suffix.lower()

    if ext == ".xlsx":
        try:
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        except Exception:
            return "unknown"
        text = " ".join(ws.title for ws in wb.worksheets).lower()
        for ws in list(wb.worksheets)[:2]:
            for row in ws.iter_rows(max_row=30, values_only=True):
                text += " " + " ".join(str(v) for v in row if v is not None).lower()
        wb.close()
        if "eventualverbindlichkeit" in text or ("leasing warehouse" in text and "car lease" in text):
            return TYPE_EVENTUALV
        return "unknown"

    if ext == ".pdf":
        try:
            with pdfplumber.open(str(path)) as pdf:
                text = " ".join((p.extract_text() or "") for p in pdf.pages[:2]).lower()
        except Exception:
            return "unknown"
        # "Steuerabgrenzung (UGB und StR)" ist spezifisch fürs Detail-PDF
        # (der Anhang selbst sagt "Latente Steuerschulden und Steueransprüche").
        if "steuerabgrenzung" in text and "leasing kfz" in text:
            return TYPE_LST_DETAIL
        return "unknown"

    return "unknown"


# ---------------------------------------------------------------------------
# Die Pipeline
# ---------------------------------------------------------------------------
class HankookPipeline(Pipeline):
    """Dokumenten-Pipeline für Hankook Tire Austria GmbH."""

    name = "hankook"

    # Standard-Zuordnungen + Hankook-Verpflichtungen + latente Steuern
    section_to_type = {
        **Pipeline.section_to_type,
        "Verpflichtungen_Leasing": TYPE_LEASING,
        "Verpflichtungen_Miete": TYPE_MIETE,
        "LatSteuer_Leasing_KFZ": TYPE_LST_LEASING,
        "LatSteuer_Firmenwert": TYPE_LST_FIRMENWERT,
        "LatSteuer_Jubilaeum": TYPE_LST_JUBILAEUM,
    }

    # ---- Modus 1: Vorjahresvergleich ----
    def extract_anhang_items(self, pdf_path: Path) -> list[AnhangItem]:
        items = super().extract_anhang_items(pdf_path)
        return [it for it in items if _is_hankook_vjv_item(it)]

    # ---- Modus 2: Belegprüfung ----
    def extract_anhang_positions(self, pdf_path: Path) -> list[AnhangPosition]:
        # Standard (Haftungen, Arbeitnehmer) + Verpflichtungen + latente Steuern
        return (
            super().extract_anhang_positions(pdf_path)
            + _extract_hankook_anhang_positions(pdf_path)
            + _extract_hankook_latente_steuern_anhang(pdf_path)
        )

    def detect_beleg_type(self, path: Path) -> str:
        t = _detect_hankook_type(path)
        return t if t != "unknown" else super().detect_beleg_type(path)

    def extract_beleg_facts(self, path: Path, dtype: str) -> list[ExtractedFact]:
        if dtype == TYPE_EVENTUALV:
            return _extract_hankook_eventualverbindlichkeiten(path)
        if dtype == TYPE_LST_DETAIL:
            return _extract_hankook_latente_steuern_beleg(path)
        return super().extract_beleg_facts(path, dtype)

    def match_tolerance(self, anhang_value) -> float:
        # Hankooks Anhang rundet auf ganze EUR. Weist der Anhang einen
        # ganzzahligen Betrag aus, ist eine Sub-Euro-Differenz zum
        # cent-genauen Beleg reine Darstellungsrundung -> OK.
        if anhang_value is not None and float(anhang_value) == round(float(anhang_value)):
            return 0.5
        return 0.02
