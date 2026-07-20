"""Checklist loading and management."""

from pathlib import Path
from typing import Optional
import json
import re

from ...models.checklist import Checklist, ChecklistItem
from ...utils.logging_config import get_logger

logger = get_logger("checklist_loader")


class ChecklistLoader:
    """
    Loads and manages audit checklists.

    Can load from:
    - JSON definition files (structured)
    - PDF files (requires parsing)
    - Manual definition

    NOTE: PDF parsing of checklists is inherently error-prone.
    For production use, consider maintaining a JSON master file
    that is manually curated from the PDF checklist.
    """

    def __init__(self, config: Optional[dict] = None):
        self.config = config or {}

    def load_from_json(self, file_path: Path) -> Checklist:
        """
        Load checklist from JSON file.

        Expected format:
        {
            "name": "PwC Anhangscheckliste",
            "version": "2024",
            "items": [
                {
                    "item_id": "item_001",
                    "category": "Allgemeine Angaben",
                    "description": "...",
                    "ugb_references": ["§ 236 Abs 1"],
                    "search_keywords": ["..."],
                    "is_mandatory": true
                }
            ]
        }
        """
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        checklist = Checklist(
            name=data.get("name", "Unnamed Checklist"),
            version=data.get("version", ""),
            source_file=str(file_path),
        )

        for item_data in data.get("items", []):
            item = ChecklistItem(
                item_id=item_data["item_id"],
                category=item_data.get("category", ""),
                description=item_data.get("description", ""),
                ugb_references=item_data.get("ugb_references", []),
                search_keywords=item_data.get("search_keywords", []),
                applicable_to=item_data.get("applicable_to", ["alle"]),
                is_mandatory=item_data.get("is_mandatory", True),
                notes=item_data.get("notes", ""),
                requires_professional_judgment=item_data.get(
                    "requires_professional_judgment", True
                ),
                judgment_guidance=item_data.get("judgment_guidance", ""),
            )
            checklist.add_item(item)

        logger.info(
            f"Loaded checklist '{checklist.name}' with {len(checklist.items)} items"
        )

        return checklist

    @staticmethod
    def _split_multi(value) -> list[str]:
        """Zerlegt 'A; B; C' / Zeilenumbrüche in eine Liste."""
        if value is None:
            return []
        parts = re.split(r"[;\n]+", str(value))
        return [p.strip() for p in parts if p.strip()]

    @staticmethod
    def _truthy(value) -> bool:
        s = str(value).strip().lower()
        return s in ("ja", "j", "yes", "y", "wahr", "true", "1", "x", "pflicht")

    def load_from_xlsx(self, file_path: Path) -> Checklist:
        """Lädt das UGB-Prüfprogramm aus einer Excel-Datei.

        Erwartete Spaltenüberschriften (Reihenfolge egal, Groß/Klein egal):
          ID | Kategorie | Prüffrage | UGB-§ | Stichwörter | Pflicht |
          Anwendbar auf | Hinweis
        Mehrere Werte (UGB-§, Stichwörter, Anwendbar) mit ';' trennen.
        So kann das Prüfprogramm jederzeit in Excel erweitert werden.
        """
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)

        # Das richtige Blatt wählen: die Excel enthält ein "Start"-Übersichts-
        # blatt, das Master-Blatt mit ALLEN Prüfpunkten und je Kategorie ein
        # Blatt. Wir nehmen das Blatt mit passender Kopfzeile (ID + Prüffrage)
        # und den MEISTEN Zeilen (= Master) – nicht stur wb.active.
        def _is_pp_header(hrow) -> bool:
            h = [str(x).strip().lower() if x is not None else "" for x in (hrow or ())]
            has_id = any(x in ("id", "nr", "item_id") for x in h)
            has_desc = any(x in ("prüffrage", "prueffrage", "frage", "beschreibung", "description") for x in h)
            return has_id and has_desc

        ws, rows = None, None
        for cand in wb.worksheets:
            crows = [r for r in cand.iter_rows(values_only=True)]
            if crows and _is_pp_header(crows[0]) and (rows is None or len(crows) > len(rows)):
                ws, rows = cand, crows
        if rows is None:                       # Fallback: bisheriges Verhalten
            ws = wb.active
            rows = [r for r in ws.iter_rows(values_only=True)]
        if not rows:
            return Checklist(name="UGB-Prüfprogramm (leer)", version="", source_file=str(file_path))

        header = [str(h).strip().lower() if h is not None else "" for h in rows[0]]

        def col(*names: str):
            for n in names:
                if n in header:
                    return header.index(n)
            return None

        ix = {
            "id":      col("id", "nr", "item_id"),
            "kat":     col("kategorie", "category"),
            "desc":    col("prüffrage", "prueffrage", "frage", "beschreibung", "description"),
            "ugb":     col("ugb-§", "ugb", "ugb-referenz", "paragraph", "§", "ugb_references"),
            "kw":      col("stichwörter", "stichworte", "keywords", "search_keywords"),
            "pflicht": col("pflicht", "is_mandatory", "mandatory"),
            "anw":     col("anwendbar auf", "anwendbar", "applicable_to"),
            "gk":      col("größenklasse (kpmg)", "größenklasse", "groessenklasse"),
            "hinweis": col("hinweis", "judgment_guidance", "guidance", "notes"),
        }

        checklist = Checklist(
            name="UGB-Prüfprogramm (Excel)",
            version="",
            source_file=str(file_path),
        )

        def cell(row, key):
            i = ix[key]
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        n = 0
        for row in rows[1:]:
            if not any(c is not None and str(c).strip() for c in row):
                continue
            desc = cell(row, "desc")
            if not desc:
                continue
            n += 1
            anw = self._split_multi(cell(row, "anw")) or ["alle"]
            checklist.add_item(ChecklistItem(
                item_id=cell(row, "id") or f"chk_{n:03d}",
                category=cell(row, "kat") or "Allgemein",
                description=desc,
                ugb_references=self._split_multi(cell(row, "ugb")),
                search_keywords=self._split_multi(cell(row, "kw")),
                applicable_to=anw,
                is_mandatory=self._truthy(cell(row, "pflicht")) if cell(row, "pflicht") else True,
                judgment_guidance=cell(row, "hinweis"),
                size_classes=self._split_multi(cell(row, "gk")),
            ))

        logger.info(f"UGB-Prüfprogramm aus Excel geladen: {len(checklist.items)} Prüfpunkte")
        return checklist

    def save_to_xlsx(self, checklist: Checklist, file_path: Path) -> None:
        """Schreibt eine Checkliste als bearbeitbares Excel-Prüfprogramm."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "UGB-Prüfprogramm"
        headers = ["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter",
                   "Pflicht", "Anwendbar auf", "Hinweis"]
        widths = [10, 26, 60, 18, 40, 9, 18, 45]
        hfill = PatternFill("solid", fgColor="305496")
        hfont = Font(bold=True, color="FFFFFF")
        for c, (h, w) in enumerate(zip(headers, widths), start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.column_dimensions[chr(64 + c)].width = w
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        for r, item in enumerate(checklist.items, start=2):
            ws.cell(row=r, column=1, value=item.item_id)
            ws.cell(row=r, column=2, value=item.category)
            ws.cell(row=r, column=3, value=item.description).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=4, value="; ".join(item.ugb_references))
            ws.cell(row=r, column=5, value="; ".join(item.search_keywords)).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=6, value="Ja" if item.is_mandatory else "Nein")
            ws.cell(row=r, column=7, value="; ".join(item.applicable_to))
            ws.cell(row=r, column=8, value=item.judgment_guidance).alignment = Alignment(wrap_text=True, vertical="top")

        file_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(file_path))
        logger.info(f"UGB-Prüfprogramm als Excel gespeichert: {file_path}")

    def load_default_checklist(self) -> Checklist:
        """
        Create a default UGB Anhang checklist.

        This provides a baseline checklist based on common UGB §§ 236-243
        disclosure requirements. Should be enhanced based on specific
        audit methodology (e.g., PwC checklist structure).

        NOTE: This is a simplified baseline. Domain experts should
        expand this with comprehensive checklist items.
        """
        checklist = Checklist(
            name="UGB Anhang Basisprüfungsprogramm",
            version="1.0",
        )

        # Define baseline items
        items = [
            # Category: Allgemeine Angaben
            ChecklistItem(
                item_id="chk_001",
                category="Allgemeine Angaben",
                description="Sind die angewandten Bilanzierungs- und Bewertungsmethoden vollständig und zutreffend dargestellt?",
                ugb_references=["§ 236 Abs 1"],
                search_keywords=[
                    "Bilanzierungsmethoden", "Bewertungsmethoden",
                    "Bewertungsgrundsätze", "Grundlagen"
                ],
                judgment_guidance="Prüfen ob alle wesentlichen Posten methodisch erläutert sind",
            ),
            ChecklistItem(
                item_id="chk_002",
                category="Allgemeine Angaben",
                description="Sind Abweichungen von Bilanzierungs- und Bewertungsmethoden gegenüber dem Vorjahr angegeben und begründet?",
                ugb_references=["§ 236 Abs 1"],
                search_keywords=[
                    "Abweichung", "Änderung", "Vorjahr", "Stetigkeit"
                ],
                judgment_guidance="Bei Methodenwechsel: Begründung und Auswirkungen prüfen",
            ),

            # Category: Anlagevermögen
            ChecklistItem(
                item_id="chk_010",
                category="Anlagevermögen",
                description="Ist die Entwicklung des Anlagevermögens (Anlagenspiegel) vollständig dargestellt?",
                ugb_references=["§ 237 Z 2"],
                search_keywords=[
                    "Anlagenspiegel", "Anlagevermögen", "Entwicklung",
                    "Zugänge", "Abgänge"
                ],
                applicable_to=["mittelgroß", "groß"],
                judgment_guidance="Abstimmung mit Bilanz, Vollständigkeit der Bewegungen",
            ),
            ChecklistItem(
                item_id="chk_011",
                category="Anlagevermögen",
                description="Sind außerplanmäßige Abschreibungen mit Begründung angegeben?",
                ugb_references=["§ 237 Z 2"],
                search_keywords=[
                    "außerplanmäßig", "Abschreibung", "Wertminderung",
                    "Wertberichtigung"
                ],
                judgment_guidance="Begründung auf Plausibilität prüfen",
            ),

            # Category: Forderungen
            ChecklistItem(
                item_id="chk_020",
                category="Forderungen und Verbindlichkeiten",
                description="Sind Forderungen mit einer Restlaufzeit von mehr als einem Jahr angegeben?",
                ugb_references=["§ 237 Z 4"],
                search_keywords=[
                    "Forderungen", "Restlaufzeit", "langfristig"
                ],
                judgment_guidance="Betrag und Art der langfristigen Forderungen prüfen",
            ),

            # Category: Verbindlichkeiten
            ChecklistItem(
                item_id="chk_030",
                category="Forderungen und Verbindlichkeiten",
                description="Sind Verbindlichkeiten nach Restlaufzeiten gegliedert?",
                ugb_references=["§ 237 Z 5"],
                search_keywords=[
                    "Verbindlichkeiten", "Restlaufzeit", "Fristigkeit"
                ],
                judgment_guidance="Gliederung: bis 1 Jahr, 1-5 Jahre, über 5 Jahre",
            ),
            ChecklistItem(
                item_id="chk_031",
                category="Forderungen und Verbindlichkeiten",
                description="Sind gesicherte Verbindlichkeiten unter Angabe von Art und Form der Sicherheiten angegeben?",
                ugb_references=["§ 237 Z 6"],
                search_keywords=[
                    "Sicherheit", "besichert", "Pfandrecht", "Hypothek"
                ],
                judgment_guidance="Art der Sicherheiten und deren Umfang prüfen",
            ),

            # Category: Rückstellungen
            ChecklistItem(
                item_id="chk_040",
                category="Rückstellungen",
                description="Sind wesentliche sonstige Rückstellungen erläutert?",
                ugb_references=["§ 237 Z 7"],
                search_keywords=[
                    "Rückstellungen", "wesentlich", "Erläuterung"
                ],
                judgment_guidance="Wesentlichkeitsgrenze anwenden, Erläuterungstiefe beurteilen",
            ),

            # Category: Haftungsverhältnisse
            ChecklistItem(
                item_id="chk_050",
                category="Haftungsverhältnisse",
                description="Sind Haftungsverhältnisse (Eventualverbindlichkeiten) vollständig angegeben?",
                ugb_references=["§ 237 Z 8", "§ 199"],
                search_keywords=[
                    "Haftungsverhältnisse", "Eventualverbindlichkeiten",
                    "Bürgschaften", "Garantien"
                ],
                judgment_guidance="Vollständigkeit kritisch hinterfragen, Managementbefragung",
            ),
            ChecklistItem(
                item_id="chk_051",
                category="Haftungsverhältnisse",
                description="Sind sonstige finanzielle Verpflichtungen angegeben?",
                ugb_references=["§ 237 Z 9"],
                search_keywords=[
                    "finanzielle Verpflichtungen", "Leasing",
                    "Mietverträge", "Bestellobligo"
                ],
                judgment_guidance="Off-Balance-Sheet Verpflichtungen erfragen",
            ),

            # Category: Umsatzerlöse
            ChecklistItem(
                item_id="chk_060",
                category="GuV-Angaben",
                description="Sind die Umsatzerlöse nach Tätigkeitsbereichen und geografischen Märkten aufgegliedert?",
                ugb_references=["§ 237 Z 10"],
                search_keywords=[
                    "Umsatzerlöse", "Aufgliederung", "Segmente"
                ],
                applicable_to=["mittelgroß", "groß"],
                judgment_guidance="Segmentierung auf Sinnhaftigkeit prüfen",
            ),

            # Category: Personal
            ChecklistItem(
                item_id="chk_070",
                category="Personalangaben",
                description="Ist die durchschnittliche Zahl der Arbeitnehmer angegeben?",
                ugb_references=["§ 237 Z 11"],
                search_keywords=[
                    "Arbeitnehmer", "Mitarbeiter", "durchschnittlich",
                    "Personalstand"
                ],
                judgment_guidance="Berechnungsmethode und Gliederung prüfen",
            ),
            ChecklistItem(
                item_id="chk_071",
                category="Personalangaben",
                description="Sind Vergütungen für Geschäftsführer/Vorstand und Aufsichtsrat angegeben?",
                ugb_references=["§ 239"],
                search_keywords=[
                    "Vergütung", "Bezüge", "Geschäftsführer",
                    "Vorstand", "Aufsichtsrat"
                ],
                applicable_to=["mittelgroß", "groß"],
                judgment_guidance="Vollständigkeit, ggf. Schutzklausel beachten",
            ),

            # Category: Organe
            ChecklistItem(
                item_id="chk_080",
                category="Organangaben",
                description="Sind die Namen der Geschäftsführer/Vorstände und Aufsichtsratsmitglieder angegeben?",
                ugb_references=["§ 241"],
                search_keywords=[
                    "Geschäftsführer", "Vorstand", "Aufsichtsrat",
                    "Namen", "Mitglieder"
                ],
                judgment_guidance="Vollständigkeit mit HR-Auszug abgleichen",
            ),

            # Category: Eigenkapital
            ChecklistItem(
                item_id="chk_090",
                category="Eigenkapital",
                description="Ist die Entwicklung der Eigenkapitalbestandteile dargestellt?",
                ugb_references=["§ 240"],
                search_keywords=[
                    "Eigenkapital", "Entwicklung", "Eigenkapitalspiegel"
                ],
                applicable_to=["mittelgroß", "groß"],
                judgment_guidance="Abstimmung mit Bilanz, Ergebnisverwendung prüfen",
            ),
        ]

        for item in items:
            checklist.add_item(item)

        logger.info(
            f"Created default checklist with {len(checklist.items)} items"
        )

        return checklist

    def save_to_json(self, checklist: Checklist, file_path: Path) -> None:
        """Save checklist to JSON file for future use."""
        data = {
            "name": checklist.name,
            "version": checklist.version,
            "items": [
                {
                    "item_id": item.item_id,
                    "category": item.category,
                    "description": item.description,
                    "ugb_references": item.ugb_references,
                    "search_keywords": item.search_keywords,
                    "applicable_to": item.applicable_to,
                    "is_mandatory": item.is_mandatory,
                    "notes": item.notes,
                    "requires_professional_judgment": item.requires_professional_judgment,
                    "judgment_guidance": item.judgment_guidance,
                }
                for item in checklist.items
            ]
        }

        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"Saved checklist to {file_path}")

    def parse_pdf_checklist(self, file_path: Path) -> Checklist:
        """
        Attempt to parse a checklist from PDF.

        WARNING: This is experimental and likely incomplete.
        PDF parsing of complex checklists is error-prone.
        Manual curation of the resulting checklist is strongly recommended.

        NOTE TO IMPLEMENTERS: This method should be enhanced based
        on the specific structure of the PwC checklist PDF.
        """
        from ...parsers.pdf_parser import PDFParser

        parser = PDFParser()
        text = parser.extract_text(file_path)

        checklist = Checklist(
            name="Parsed Checklist (requires manual review)",
            version="auto-parsed",
            source_file=str(file_path),
        )

        # Basic heuristic parsing - looks for numbered items with UGB references
        # Pattern: number. Description text (§ xxx)
        item_pattern = re.compile(
            r"(\d+(?:\.\d+)?)\.\s+(.+?)(?:\(([§\s\d,]+)\))?(?:\n|$)",
            re.MULTILINE
        )

        current_category = "Allgemein"
        item_count = 0

        for match in item_pattern.finditer(text):
            number = match.group(1)
            description = match.group(2).strip()
            ugb_refs = match.group(3)

            if len(description) < 10:
                continue

            # Parse UGB references
            references = []
            if ugb_refs:
                ref_matches = re.findall(r"§\s*\d+[a-z]?", ugb_refs)
                references = list(set(ref_matches))

            item_count += 1
            item = ChecklistItem(
                item_id=f"parsed_{item_count:03d}",
                category=current_category,
                description=description,
                ugb_references=references,
                search_keywords=self._extract_keywords(description),
            )
            checklist.add_item(item)

        logger.warning(
            f"Parsed {len(checklist.items)} items from PDF. "
            "Manual review required!"
        )

        return checklist

    def _extract_keywords(self, text: str) -> list[str]:
        """Extract potential keywords from description text."""
        # Remove common stop words and short words
        stop_words = {
            "der", "die", "das", "und", "oder", "mit", "für", "von",
            "ist", "sind", "werden", "wurde", "hat", "haben", "wird",
            "bei", "als", "auch", "auf", "aus", "dem", "den", "des",
            "ein", "eine", "einer", "einem", "einen", "sich", "nicht",
            "nach", "über", "unter", "vor", "zum", "zur"
        }

        words = re.findall(r"\b[A-ZÄÖÜa-zäöüß]{4,}\b", text)
        keywords = [
            w for w in words
            if w.lower() not in stop_words
        ]

        # Return unique keywords, max 5
        seen = set()
        unique = []
        for kw in keywords:
            if kw.lower() not in seen:
                seen.add(kw.lower())
                unique.append(kw)
                if len(unique) >= 5:
                    break

        return unique
