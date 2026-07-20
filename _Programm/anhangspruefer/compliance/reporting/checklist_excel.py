"""
Ausgefüllte KPMG-Checkliste als Excel-Arbeitspapier (Modus 3).

Statt eines Fließtext-Protokolls wird die KPMG-Prüfliste selbst befüllt:
je Prüfpunkt Ergebnis, Fundstelle/Nachweis und Anmerkung. So entsteht ein
Arbeitspapier, das wie eine von Hand ausgefüllte Checkliste aussieht.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ...models.checklist import Checklist
from ...models.finding import ReviewResult


# Klare Verdikte: Ja / Fehlt / n. a. — "Offen" nur, wenn automatisch nicht
# entscheidbar (manuell zu prüfen). Kein "Teilweise".
_STATUS_LABEL = {
    "ENTSPRICHT": "Ja",
    "TEILWEISE ENTSPRECHEND": "Offen",
    "NICHT ENTSPRECHEND": "Fehlt",
    "NICHT BEURTEILBAR": "Offen",
    "NICHT ANWENDBAR": "n. a.",
    "PRÜFUNG AUSSTEHEND": "Offen",
}
_STATUS_FILL = {
    "ENTSPRICHT": "C6EFCE",
    "TEILWEISE ENTSPRECHEND": "FFEB9C",
    "NICHT ENTSPRECHEND": "FFC7CE",
    "NICHT BEURTEILBAR": "FFEB9C",
    "NICHT ANWENDBAR": "E7E6E6",
    "PRÜFUNG AUSSTEHEND": "FFEB9C",
}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_TOP = Alignment(wrap_text=True, vertical="top")
_CTR = Alignment(horizontal="center", vertical="top")


def _evidence_text(finding) -> str:
    """EINE Fundstelle (Seite + kurzes Zitat). Weniger ist mehr."""
    if finding is None or not finding.evidence:
        return ""
    ev = finding.evidence[0]
    quote = " ".join((ev.quote or "").split())
    if len(quote) > 220:
        quote = quote[:220] + "…"
    ref = f"S. {ev.page_number}: " if ev.page_number else ""
    return (ref + quote) if quote else ""


def _note_text(finding, status: str, item) -> str:
    """Kurze, klare Anmerkung – eine Aussage, kein Fließtext."""
    reasoning = (finding.technical_reasoning if finding else "").strip()
    if status == "NICHT ANWENDBAR":
        return reasoning                     # kurzer n.-a.-Grund aus dem Filter
    if status == "NICHT ENTSPRECHEND":
        # KONKRET sagen, was fehlt (aus KI bzw. Fallback: die Prüffrage).
        # Generische Matcher-Floskeln ("automatisch identifiziert") verwerfen.
        was = ""
        for m in (finding.missing_elements if finding else None) or []:
            m = (m or "").strip()
            if m and "automatisch identifiziert" not in m.lower():
                was = m
                break
        was = was or item.description.strip()[:160]
        return f"Fehlt: {was}"
    if reasoning.startswith("KI:") and status == "NICHT BEURTEILBAR":
        return reasoning                     # warum unklar -> manuell prüfen
    return ""                                # Ja/Offen: keine Textwüste


def generate_checklist_xlsx(checklist: Checklist, result: ReviewResult, out_path: Path) -> None:
    by_id = {f.checklist_item_id: f for f in result.findings}
    wb = openpyxl.Workbook()

    # -------- Blatt 1: Übersicht --------
    ov = wb.active
    ov.title = "Übersicht"
    ov.cell(row=1, column=1, value="UGB-Anhang – ausgefüllte KPMG-Checkliste").font = Font(bold=True, size=14)
    ov.cell(row=2, column=1, value=f"Dokument: {result.document_name}")
    ov.cell(row=3, column=1, value=f"Prüfprogramm: {result.checklist_name}")

    counts: dict[str, int] = {}
    for f in result.findings:
        counts[f.status.value] = counts.get(f.status.value, 0) + 1
    ov.cell(row=5, column=1, value="Ergebnis").font = Font(bold=True)
    ov.cell(row=5, column=2, value="Anzahl").font = Font(bold=True)
    order = ["ENTSPRICHT", "TEILWEISE ENTSPRECHEND", "NICHT ENTSPRECHEND",
             "NICHT BEURTEILBAR", "NICHT ANWENDBAR", "PRÜFUNG AUSSTEHEND"]
    r = 6
    for key in order:
        if counts.get(key):
            ov.cell(row=r, column=1, value=_STATUS_LABEL.get(key, key))
            ov.cell(row=r, column=2, value=counts[key])
            ov.cell(row=r, column=1).fill = PatternFill("solid", fgColor=_STATUS_FILL.get(key, "FFFFFF"))
            r += 1
    ov.cell(row=r + 1, column=1, value="Geprüfte Punkte gesamt").font = Font(bold=True)
    ov.cell(row=r + 1, column=2, value=len(result.findings)).font = Font(bold=True)

    # Final-Zählung (inkl. Prüfer-Übersteuerung) – live per Formel
    last = len(checklist.items) + 1
    fr = r + 3
    ov.cell(row=fr, column=1, value="Final (inkl. Prüfer-Übersteuerung)").font = Font(bold=True)
    for off, verdict in enumerate(("Ja", "Fehlt", "n. a.", "Offen"), start=1):
        ov.cell(row=fr + off, column=1, value=verdict)
        ov.cell(row=fr + off, column=2,
                value=f"=COUNTIF('KPMG-Checkliste (ausgefüllt)'!$K$2:$K${last},\"{verdict}\")")

    ov.cell(row=fr + 6, column=1,
            value=("'n. a.' = Position/Sachverhalt liegt nicht vor. 'Offen' = automatisch "
                   "nicht entscheidbar, manuell prüfen. Prüfer-Urteil (Spalte J) übersteuert "
                   "das Tool-Ergebnis; Spalte K zeigt das maßgebliche Final-Urteil."
                   )).font = Font(italic=True, color="7F7F7F")
    ov.column_dimensions["A"].width = 42
    ov.column_dimensions["B"].width = 12

    # -------- Blatt 2: Ausgefüllte Checkliste --------
    # Übersteuerung durch den Prüfer: Spalte "Prüfer-Urteil" (Dropdown) schlägt
    # das Tool-Ergebnis; Spalte "Final" zeigt das maßgebliche Urteil.
    ws = wb.create_sheet("KPMG-Checkliste (ausgefüllt)")
    headers = ["Nr.", "ID", "Kategorie", "Prüffrage (KPMG)", "UGB-§ / Fachgutachten",
               "Ergebnis (Tool)", "Nachweis / Fundstelle im Anhang", "Anmerkung",
               "Pflicht", "Prüfer-Urteil", "Final", "Prüfer-Kommentar"]
    widths = [5, 7, 24, 58, 26, 14, 50, 40, 8, 13, 10, 36]
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CTR
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(c)].width = w

    for i, item in enumerate(checklist.items, start=1):
        f = by_id.get(item.item_id)
        status = f.status.value if f else "PRÜFUNG AUSSTEHEND"
        r_idx = i + 1
        values = [
            i,
            item.item_id,
            item.category,
            item.description,
            "; ".join(item.ugb_references),
            _STATUS_LABEL.get(status, status),
            "" if status == "NICHT ANWENDBAR" else _evidence_text(f),
            _note_text(f, status, item),
            "Ja" if item.is_mandatory else "Nein",
            "",                                            # Prüfer-Urteil (Dropdown)
            f'=IF(J{r_idx}<>"",J{r_idx},F{r_idx})',        # Final = Prüfer schlägt Tool
            "",                                            # Prüfer-Kommentar
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.alignment = _CTR if c in (1, 2, 6, 9, 10, 11) else _TOP
            cell.border = _BORDER
        ws.cell(row=r_idx, column=6).fill = PatternFill(
            "solid", fgColor=_STATUS_FILL.get(status, "FFFFFF"))

    n_rows = len(checklist.items) + 1

    # Dropdown für das Prüfer-Urteil (leer = Tool-Ergebnis gilt)
    dv = DataValidation(type="list", formula1='"Ja,Fehlt,n. a.,Offen"', allow_blank=True,
                        promptTitle="Prüfer-Urteil",
                        prompt="Leer lassen = Tool-Ergebnis gilt. Auswahl übersteuert das Tool.")
    ws.add_data_validation(dv)
    dv.add(f"J2:J{n_rows}")

    # Final-Spalte farblich nach Urteil (bedingte Formatierung)
    _final_colors = {"Ja": "C6EFCE", "Fehlt": "FFC7CE", "n. a.": "E7E6E6", "Offen": "FFEB9C"}
    for verdict, color in _final_colors.items():
        ws.conditional_formatting.add(
            f"K2:K{n_rows}",
            CellIsRule(operator="equal", formula=[f'"{verdict}"'],
                       fill=PatternFill("solid", fgColor=color)))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_rows}"

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
