"""
Excel-Export des Vorjahresvergleichs.

Blätter:
  1. Übersicht                 — Zahlenstatistik + Textzähler
  2. Alle Posten               — Vollständige Zahlenliste (Bilanz, GuV, Anhang)
  3. Abweichungen              — Nur Status ABWEICHUNG
  4. Nur aktuell               — Posten nur im neuen Abschluss
  5. Nur Vorjahr               — Posten nur im Vorjahresabschluss
  6. Neu im Bericht            — Text, der im Vorjahr fehlt
  7. Fehlt gegenüber Vorjahr   — Text, der heuer fehlt
  8. Geänderter Text           — gleicher Absatz, anderer Wortlaut
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .comparator import CompareResult
from .text_compare import diff_excerpt


STATUS_FILLS = {
    "OK":             PatternFill("solid", fgColor="C6EFCE"),
    "ABWEICHUNG":     PatternFill("solid", fgColor="FFC7CE"),
    "NUR_AKTUELL":    PatternFill("solid", fgColor="FFEB9C"),
    "NUR_VORJAHR":    PatternFill("solid", fgColor="FFEB9C"),
    "FEHLENDER_WERT": PatternFill("solid", fgColor="D9D9D9"),
}

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HINT_FONT = Font(italic=True, color="7F7F7F")
TOP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="top")


def _set_header(ws, row, headers, widths):
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        c.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _write_uebersicht(wb: Workbook, result: CompareResult) -> None:
    ws = wb.active
    ws.title = "Übersicht"

    title = ws.cell(row=1, column=1, value="Vorjahresvergleich – Bilanz, GuV, Anhang")
    title.font = Font(bold=True, size=14)
    ws.merge_cells("A1:D1")

    ws.cell(row=3, column=1, value="Aktueller Anhang:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=result.current_pdf.name)
    ws.cell(row=4, column=1, value="Vorjahres-Anhang:").font = Font(bold=True)
    ws.cell(row=4, column=2, value=result.prior_pdf.name)
    ws.merge_cells("B3:D3")
    ws.merge_cells("B4:D4")

    ws.cell(row=6, column=1, value="Status").font = Font(bold=True)
    ws.cell(row=6, column=2, value="Anzahl").font = Font(bold=True)

    stats = result.stats
    rows = [
        ("OK (Zahlen stimmen)",        stats.get("OK", 0)),
        ("ABWEICHUNG",                 stats.get("ABWEICHUNG", 0)),
        ("Nur im aktuellen Anhang",    stats.get("NUR_AKTUELL", 0)),
        ("Nur im Vorjahres-Anhang",    stats.get("NUR_VORJAHR", 0)),
        ("Wert fehlt auf einer Seite", stats.get("FEHLENDER_WERT", 0)),
        ("Gesamt",                      stats.get("GESAMT", 0)),
    ]
    status_keys = ["OK", "ABWEICHUNG", "NUR_AKTUELL", "NUR_VORJAHR", "FEHLENDER_WERT", None]
    for i, ((label, count), key) in enumerate(zip(rows, status_keys), start=7):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=count)
        if key in STATUS_FILLS:
            c1.fill = STATUS_FILLS[key]
            c2.fill = STATUS_FILLS[key]
        if label == "Gesamt":
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36

    fehlt = sum(1 for t in result.text_rows if t.status == "FEHLT")
    neu = sum(1 for t in result.text_rows if t.status == "NEU")
    geaendert = sum(1 for t in result.text_rows if t.status == "GEÄNDERT")
    ident = sum(1 for t in result.text_rows if t.status == "IDENT")

    ws.cell(row=14, column=1, value="Textvergleich").font = Font(bold=True)

    r15a = ws.cell(row=15, column=1, value="Neu im aktuellen Bericht")
    r15b = ws.cell(row=15, column=2, value=neu)
    r15a.fill = _TEXT_STATUS_FILL["NEU"]
    r15b.fill = _TEXT_STATUS_FILL["NEU"]
    ws.cell(row=15, column=3, value="→ Blatt 'Neu im Bericht'").font = HINT_FONT

    r16a = ws.cell(row=16, column=1, value="Fehlt gegenüber Vorjahr")
    r16b = ws.cell(row=16, column=2, value=fehlt)
    r16a.fill = _TEXT_STATUS_FILL["FEHLT"]
    r16b.fill = _TEXT_STATUS_FILL["FEHLT"]
    ws.cell(row=16, column=3, value="→ Blatt 'Fehlt gegenüber Vorjahr'").font = HINT_FONT

    r17a = ws.cell(row=17, column=1, value="Geänderter Wortlaut")
    r17b = ws.cell(row=17, column=2, value=geaendert)
    r17a.fill = _TEXT_STATUS_FILL["GEÄNDERT"]
    r17b.fill = _TEXT_STATUS_FILL["GEÄNDERT"]
    ws.cell(row=17, column=3, value="→ Blatt 'Geänderter Text'").font = HINT_FONT

    ws.cell(row=18, column=1, value="Unverändert (nicht einzeln ausgewiesen)")
    ws.cell(row=18, column=2, value=ident)

    ws.cell(row=20, column=1,
            value=("Hinweis: Heuristische Analyse (Regex + Fuzzy-Label-Matching, "
                   "Textvergleich ohne Zahlen). Ersetzt KEINE prüferische Beurteilung. "
                   "Manuelle Validierung erforderlich.")
            ).font = HINT_FONT
    ws.merge_cells("A20:D20")


def _write_rows(ws, rows, headers, widths):
    _set_header(ws, 1, headers, widths)
    last_label = None
    for r_idx, row in enumerate(rows, start=2):
        # Label nur einmal je Item-Block anzeigen (mehrspaltige Tabellen kompakter)
        label_text = row.label_current_doc or row.label_prior_doc or ""
        display_label = label_text if label_text != last_label else ""
        last_label = label_text

        ws.cell(row=r_idx, column=1, value=display_label)
        ws.cell(row=r_idx, column=2, value=getattr(row, "column_index", 1))
        c3 = ws.cell(row=r_idx, column=3, value=row.value_in_current_anhang)
        c4 = ws.cell(row=r_idx, column=4, value=row.value_in_prior_anhang)
        c5 = ws.cell(row=r_idx, column=5, value=row.difference)
        ws.cell(row=r_idx, column=6, value=row.page_current or None)
        ws.cell(row=r_idx, column=7, value=row.page_prior or None)
        c8 = ws.cell(row=r_idx, column=8, value=row.status)
        ws.cell(row=r_idx, column=9, value=round(row.match_score, 2))

        fmt = '#,##0.00'
        c3.number_format = fmt
        c4.number_format = fmt
        c5.number_format = fmt

        fill = STATUS_FILLS.get(row.status)
        if fill:
            for col in range(1, 10):
                ws.cell(row=r_idx, column=col).fill = fill

        c8.alignment = Alignment(horizontal="center")


_HEADERS = [
    "Bezeichnung",
    "Spalte",
    "Wert lt. aktuellem Anhang (Vorjahresspalte)",
    "Wert lt. Vorjahres-Anhang (Berichtsjahr)",
    "Differenz",
    "Seite aktuell",
    "Seite Vorjahr",
    "Status",
    "Match-Score",
]
_WIDTHS = [55, 8, 30, 30, 16, 10, 10, 16, 12]


def _write_alle(wb: Workbook, result: CompareResult) -> None:
    ws = wb.create_sheet("Alle Posten")
    _write_rows(ws, result.rows, _HEADERS, _WIDTHS)


_TEXT_STATUS_FILL = {
    "IDENT":    PatternFill("solid", fgColor="FFFFFF"),
    "GEÄNDERT": PatternFill("solid", fgColor="FFEB9C"),
    "NEU":      PatternFill("solid", fgColor="DDEBF7"),
    "FEHLT":    PatternFill("solid", fgColor="FFC7CE"),
}


def _write_abweichungen(wb: Workbook, result: CompareResult) -> None:
    ws = wb.create_sheet("Abweichungen")
    rows = [r for r in result.rows if r.status == "ABWEICHUNG"]
    _write_rows(ws, rows, _HEADERS, _WIDTHS)


def _write_nur_aktuell(wb: Workbook, result: CompareResult) -> None:
    ws = wb.create_sheet("Nur aktuell")
    rows = [r for r in result.rows if r.status == "NUR_AKTUELL"]
    _write_rows(ws, rows, _HEADERS, _WIDTHS)


def _write_nur_vorjahr(wb: Workbook, result: CompareResult) -> None:
    ws = wb.create_sheet("Nur Vorjahr")
    rows = [r for r in result.rows if r.status == "NUR_VORJAHR"]
    _write_rows(ws, rows, _HEADERS, _WIDTHS)


def _split_diff(current: str, prior: str) -> tuple[str, str]:
    """Zerlegt den Wort-Diff in (nur Vorjahr / entfernt, nur aktuell / neu)."""
    excerpt = diff_excerpt(current, prior, max_len=800)
    removed = added = ""
    if "Vorjahr:" in excerpt or "aktuell:" in excerpt:
        for p in excerpt.split(" || "):
            if p.startswith("Vorjahr:"):
                removed = p[len("Vorjahr:"):].strip()
            elif p.startswith("aktuell:"):
                added = p[len("aktuell:"):].strip()
    elif excerpt.startswith("nur"):
        if "Vorjahr" in excerpt:
            removed = excerpt
        else:
            added = excerpt
    else:
        added = excerpt
    return removed, added


def _write_empty(ws, message: str, cols: int) -> None:
    c = ws.cell(row=3, column=1, value=message)
    c.font = HINT_FONT
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=cols)


def _write_neu(wb: Workbook, result: CompareResult) -> None:
    """Absätze, die im aktuellen Bericht stehen und im Vorjahr fehlen."""
    ws = wb.create_sheet("Neu im Bericht")
    hint = ws.cell(
        row=1, column=1,
        value="Neu: steht im aktuellen Bericht, im Vorjahresbericht so nicht enthalten.",
    )
    hint.font = Font(bold=True)
    ws.merge_cells("A1:D1")
    headers = ["Seite aktuell", "Neuer Text"]
    widths = [14, 110]
    _set_header(ws, 2, headers, widths)
    rows = [t for t in result.text_rows if t.status == "NEU"]
    if not rows:
        _write_empty(ws, "Nichts Neues gegenüber dem Vorjahr.", 2)
        return
    for r_idx, tr in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=tr.page_current).alignment = CTR
        c2 = ws.cell(row=r_idx, column=2, value=tr.current)
        c2.alignment = TOP
        for col in range(1, 3):
            ws.cell(row=r_idx, column=col).fill = _TEXT_STATUS_FILL["NEU"]
        ws.row_dimensions[r_idx].height = min(90, 18 + 12 * (1 + len(tr.current) // 90))


def _write_fehlt(wb: Workbook, result: CompareResult) -> None:
    """Absätze, die im Vorjahr standen und heuer fehlen."""
    ws = wb.create_sheet("Fehlt gegenüber Vorjahr")
    hint = ws.cell(
        row=1, column=1,
        value="Fehlt: stand im Vorjahresbericht, im aktuellen Bericht so nicht enthalten.",
    )
    hint.font = Font(bold=True)
    ws.merge_cells("A1:D1")
    headers = ["Seite Vorjahr", "Fehlender Text (Vorjahr)"]
    widths = [14, 110]
    _set_header(ws, 2, headers, widths)
    rows = [t for t in result.text_rows if t.status == "FEHLT"]
    if not rows:
        _write_empty(ws, "Gegenüber dem Vorjahr fehlt kein Textabschnitt.", 2)
        return
    for r_idx, tr in enumerate(rows, start=3):
        ws.cell(row=r_idx, column=1, value=tr.page_prior).alignment = CTR
        c2 = ws.cell(row=r_idx, column=2, value=tr.prior)
        c2.alignment = TOP
        for col in range(1, 3):
            ws.cell(row=r_idx, column=col).fill = _TEXT_STATUS_FILL["FEHLT"]
        ws.row_dimensions[r_idx].height = min(90, 18 + 12 * (1 + len(tr.prior) // 90))


def _write_geaendert(wb: Workbook, result: CompareResult) -> None:
    """Absätze, die in beiden Berichten stehen, aber anders formuliert sind."""
    ws = wb.create_sheet("Geänderter Text")
    hint = ws.cell(
        row=1, column=1,
        value="Geändert: derselbe Absatz in beiden Berichten, abweichender Wortlaut.",
    )
    hint.font = Font(bold=True)
    ws.merge_cells("A1:F1")
    headers = [
        "S. akt.",
        "S. VJ",
        "Entfernt (nur Vorjahr)",
        "Hinzugefügt (nur aktuell)",
        "Text aktuell",
        "Text Vorjahr",
    ]
    widths = [10, 10, 40, 40, 50, 50]
    _set_header(ws, 2, headers, widths)
    rows = [t for t in result.text_rows if t.status == "GEÄNDERT"]
    if not rows:
        _write_empty(ws, "Kein geänderter Wortlaut gegenüber dem Vorjahr.", 6)
        return
    for r_idx, tr in enumerate(rows, start=3):
        removed, added = _split_diff(tr.current, tr.prior)
        ws.cell(row=r_idx, column=1, value=tr.page_current).alignment = CTR
        ws.cell(row=r_idx, column=2, value=tr.page_prior).alignment = CTR
        ws.cell(row=r_idx, column=3, value=removed).alignment = TOP
        ws.cell(row=r_idx, column=4, value=added).alignment = TOP
        ws.cell(row=r_idx, column=5, value=tr.current).alignment = TOP
        ws.cell(row=r_idx, column=6, value=tr.prior).alignment = TOP
        fill = _TEXT_STATUS_FILL["GEÄNDERT"]
        for col in range(1, 7):
            ws.cell(row=r_idx, column=col).fill = fill
        longest = max(len(tr.current or ""), len(tr.prior or ""), 1)
        ws.row_dimensions[r_idx].height = min(110, 18 + 12 * (1 + longest // 80))


def generate_excel(result: CompareResult, output_path: Path) -> Path:
    """Schreibt einen Excel-Bericht für das Vergleichsergebnis."""
    wb = Workbook()
    _write_uebersicht(wb, result)
    _write_alle(wb, result)
    _write_abweichungen(wb, result)
    _write_nur_aktuell(wb, result)
    _write_nur_vorjahr(wb, result)
    _write_neu(wb, result)
    _write_fehlt(wb, result)
    _write_geaendert(wb, result)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
