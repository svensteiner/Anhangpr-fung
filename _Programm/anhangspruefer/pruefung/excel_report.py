"""
Excel-Report für die Anhang-Prüfung gegen Detailunterlagen.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .comparator import PruefResult, STATUS_OK, STATUS_ABWEICHUNG, STATUS_HINWEIS, STATUS_KEIN_BELEG


# ---------------------------------------------------------------------------
# Farben (LLP-Palette)
# ---------------------------------------------------------------------------
C_MAROON    = "7B1818"
C_ORANGE    = "E07A1E"
C_GOLD      = "C8900A"
C_WHITE     = "FFFFFF"
C_LIGHT     = "FDF8F3"
C_GREEN     = "2E7D32"
C_RED       = "C62828"
C_YELLOW    = "F57F17"
C_GRAY      = "F5F5F5"
C_BORDER    = "D0C5B8"

_thin = Side(style="thin", color=C_BORDER)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

STATUS_COLOR = {
    STATUS_OK:         ("D4EDDA", C_GREEN),
    STATUS_HINWEIS:    ("FFF9C4", C_GOLD),
    STATUS_ABWEICHUNG: ("FFCDD2", C_RED),
    STATUS_KEIN_BELEG: ("F5F5F5", "888888"),
    "KEIN_WERT":       ("F5F5F5", "888888"),
}
STATUS_LABEL = {
    STATUS_OK:         "OK",
    STATUS_HINWEIS:    "OK (Hinweis)",
    STATUS_ABWEICHUNG: "ABWEICHUNG",
    STATUS_KEIN_BELEG: "Kein Beleg",
    "KEIN_WERT":       "Kein Wert",
}


def _hdr_cell(ws, row, col, value, bg=C_MAROON, fg=C_WHITE, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    c.border = _border
    return c


def _data_cell(ws, row, col, value, bg=None, bold=False, fmt=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=10)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=True
    )
    if fmt:
        c.number_format = fmt
    c.border = _border
    return c


def generate_excel(result: PruefResult, output_path: Path) -> Path:
    output_path = Path(output_path)
    wb = Workbook()

    _write_uebersicht(wb, result)
    _write_detail(wb, result)

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# Blatt 1: Übersicht
# ---------------------------------------------------------------------------
def _write_uebersicht(wb: Workbook, result: PruefResult) -> None:
    ws = wb.active
    ws.title = "Übersicht"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 20

    r = 1
    # Titelzeile
    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1, value="LLP · Anhangsprüfer – Prüfung gegen Detailunterlagen")
    c.font = Font(bold=True, size=14, color=C_MAROON)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 30
    r += 1

    ws.merge_cells(f"A{r}:B{r}")
    c = ws.cell(row=r, column=1, value="HAFTUNGSAUSSCHLUSS: Prüfungsunterstützung – keine Ersetzung fachlicher Beurteilung")
    c.font = Font(italic=True, size=9, color="AA6600")
    c.fill = PatternFill("solid", fgColor="FFFBF0")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 18
    r += 2

    for label, value in [
        ("Anhang-Datei", result.anhang_filename),
        ("Belege", ", ".join(result.beleg_filenames) or "—"),
        ("Geprüfte Positionen", len(result.rows)),
        ("OK / OK mit Hinweis", result.count_ok),
        ("Abweichungen", result.count_abweichung),
        ("Kein Beleg vorhanden", result.count_kein_beleg),
    ]:
        _data_cell(ws, r, 1, label, bold=True)
        _data_cell(ws, r, 2, value, align="right")
        r += 1

    r += 1
    # Status-Legende
    _hdr_cell(ws, r, 1, "Status-Legende", bg=C_MAROON)
    _hdr_cell(ws, r, 2, "Bedeutung", bg=C_MAROON)
    r += 1
    for st, (bg, fg) in STATUS_COLOR.items():
        _data_cell(ws, r, 1, STATUS_LABEL[st], bg=bg, bold=True)
        desc = {
            STATUS_OK: "Wert im Anhang stimmt mit Beleg überein",
            STATUS_HINWEIS: "Wert stimmt, aber methodischer Hinweis beachten",
            STATUS_ABWEICHUNG: "Wert im Anhang weicht vom Beleg ab",
            STATUS_KEIN_BELEG: "Kein passender Beleg hochgeladen",
            "KEIN_WERT": "Beleg vorhanden, aber kein Wert extrahierbar",
        }.get(st, "")
        _data_cell(ws, r, 2, desc)
        r += 1


# ---------------------------------------------------------------------------
# Blatt 2: Detail
# ---------------------------------------------------------------------------
def _write_detail(wb: Workbook, result: PruefResult) -> None:
    ws = wb.create_sheet("Prüfungsergebnis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    cols = [
        ("Abschnitt",            22),
        ("Position laut Anhang", 38),
        ("Wert laut Anhang",     18),
        ("Wert laut Beleg(en)",  18),
        ("Differenz",            14),
        ("Seite Anhang",         12),
        ("Belegdatei(en)",       35),
        ("Status",               16),
        ("Hinweis",              42),
    ]

    for ci, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Titelzeile
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    tc = ws.cell(row=1, column=1, value="Anhangsprüfung – Detailergebnis")
    tc.font = Font(bold=True, size=12, color=C_WHITE)
    tc.fill = PatternFill("solid", fgColor=C_MAROON)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # Header
    for ci, (hdr, _) in enumerate(cols, 1):
        _hdr_cell(ws, 2, ci, hdr, bg=C_ORANGE)
    ws.row_dimensions[2].height = 20

    # Datenzeilen
    FMT_EUR = '#,##0.00 "EUR"'
    FMT_INT = '#,##0'
    r = 3
    for row in result.rows:
        bg, _ = STATUS_COLOR.get(row.status, ("FFFFFF", "000000"))
        is_currency = row.section == "Haftungsverhaeltnisse"
        num_fmt = FMT_EUR if is_currency else FMT_INT

        belegfiles = ", ".join({f.filename for f in row.belege}) if row.belege else "—"
        belegvals  = " + ".join(
            f"{f.filename}: {f.value:,.2f}" for f in row.belege
        ) if row.belege else ""

        vals = [
            row.section.replace("verhaeltnisse", "verhältnisse"),
            row.label,
            row.anhang_value,
            row.beleg_value,
            row.difference,
            row.page_anhang,
            belegfiles,
            STATUS_LABEL.get(row.status, row.status),
            row.note or "",
        ]
        fmts = [None, None, num_fmt, num_fmt, num_fmt, None, None, None, None]
        aligns = ["left", "left", "right", "right", "right", "center", "left", "center", "left"]

        for ci, (val, fmt, align) in enumerate(zip(vals, fmts, aligns), 1):
            _data_cell(ws, r, ci, val, bg=bg, fmt=fmt, align=align)

        # Status-Zelle farbig
        st_cell = ws.cell(row=r, column=8)
        sbg, sfg = STATUS_COLOR.get(row.status, ("FFFFFF", "000000"))
        st_cell.fill = PatternFill("solid", fgColor=sbg)
        st_cell.font = Font(bold=True, color=sfg, size=10)

        ws.row_dimensions[r].height = 28 if row.note else 20
        r += 1
