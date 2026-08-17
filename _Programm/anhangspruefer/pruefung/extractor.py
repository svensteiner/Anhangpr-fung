"""
Extraktoren für die Anhang-Prüfung gegen Detailunterlagen.

Unterstützte Belegtypen (automatische Erkennung):
  - Bankgarantien / Haftungsobligo  (PDF)
  - Personalstand / Mitarbeiter     (Excel .xlsx)
  - Forderungenspiegel              (PDF)
  - Rückstellungsspiegel            (PDF)
  - Verbindlichkeitenspiegel        (PDF)

Aus dem Anhang werden extrahiert:
  - Haftungsverhältnisse  (Gesamtbetrag)
  - Arbeitnehmer          (durchschnittlich, Gesamt)
  - Summe Forderungen     (Gesamtbetrag)
  - Summe Rückstellungen  (Stand Berichtsjahr)
  - Summe Verbindlichkeiten (Gesamtbetrag)
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pdfplumber
import openpyxl


# ---------------------------------------------------------------------------
# Zahlenerkennung (deutsches Format)
# ---------------------------------------------------------------------------
_NUMBER_RE = re.compile(
    r"""
    (?<![\w.,-])
    \(?-?
    (?:
        \d{1,3}(?:\.\d{3})+(?:,\d+)?
        |
        \d+(?:,\d+)?
    )
    \)?
    (?![\w.])
    """,
    re.VERBOSE,
)


def _parse_de_number(s: str) -> Optional[float]:
    s = s.strip()
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg, s = True, s[1:-1]
    if s.startswith("-"):
        neg, s = True, s[1:]
    s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return None


def _trailing_numbers(line: str) -> list[float]:
    """Alle Zahlen am Zeilenende (kontinuierlich, kein Text dazwischen)."""
    ms = list(_NUMBER_RE.finditer(line))
    if not ms:
        return []
    nums, label_end = [], len(line)
    for m in reversed(ms):
        if line[m.end():label_end].strip():
            break
        v = _parse_de_number(m.group())
        if v is None:
            break
        nums.insert(0, v)
        label_end = m.start()
    return nums


def _all_numbers(line: str) -> list[float]:
    """Alle Zahlen in der Zeile."""
    return [v for m in _NUMBER_RE.finditer(line) if (v := _parse_de_number(m.group())) is not None]


# ---------------------------------------------------------------------------
# Datenklassen
# ---------------------------------------------------------------------------
@dataclass
class AnhangPosition:
    """Eine aus dem Anhang extrahierte Prüfposition."""
    section: str           # z.B. "Haftungsverhaeltnisse"
    label: str             # menschenlesbare Bezeichnung
    current_value: Optional[float]   # Berichtsjahr (linke Spalte)
    prior_value: Optional[float]     # Vorjahr (rechte Spalte)
    page: int
    source_line: str = ""
    note: str = ""         # optionaler Hinweis (z.B. Methodik-Unterschied)


@dataclass
class ExtractedFact:
    """Ein aus einem Beleg extrahiertes Datum."""
    source_type: str       # "bank_guarantees" | "hr_employees" | ...
    label: str
    value: float
    filename: str = ""
    source_line: str = ""
    sheet: str = ""


# ---------------------------------------------------------------------------
# Dokumenttyp-Erkennung
# ---------------------------------------------------------------------------
def detect_type(path: Path) -> str:
    """
    Erkennt automatisch den Dokumenttyp anhand von Schlüsselwörtern.
    Gibt zurück: "bank_guarantees" | "hr_employees" | "anhang" |
                 "forderungen_spiegel" | "rueckstellungen_spiegel" |
                 "verbindlichkeiten_spiegel" | "unknown"
    """
    path = Path(path)
    ext = path.suffix.lower()

    if ext == ".xlsx":
        return _detect_excel_type(path)

    if ext == ".pdf":
        return _detect_pdf_type(path)

    return "unknown"


def _detect_pdf_type(path: Path) -> str:
    try:
        with pdfplumber.open(str(path)) as pdf:
            # Erste 2 Seiten reichen zur Erkennung
            text = ""
            for page in pdf.pages[:2]:
                text += (page.extract_text() or "") + "\n"
    except Exception:
        return "unknown"

    text_lower = text.lower()

    # Spiegel-PDFs zuerst (sehr spezifisch, höhere Priorität als Anhang)
    if "forderungenspiegel" in text_lower or "summe forderungen" in text_lower:
        return "forderungen_spiegel"
    if "rückstellungsspiegel" in text_lower or "ruecksstellungsspiegel" in text_lower \
            or "rueckstellungsspiegel" in text_lower \
            or "summe rückstellungen" in text_lower or "summe rueckstellungen" in text_lower:
        return "rueckstellungen_spiegel"
    if "verbindlichkeitenspiegel" in text_lower or "summe verbindlichkeiten" in text_lower:
        return "verbindlichkeiten_spiegel"

    # Bankgarantie / Haftungsobligo
    if any(kw in text_lower for kw in [
        "haftungsobligo", "obligo", "garantieart", "mietgarantie",
        "haftrücklassgarantie", "bankgarantie", "eurggw",
    ]):
        return "bank_guarantees"

    # Anhang / Jahresabschluss
    if any(kw in text_lower for kw in [
        "bilanzierungs", "bewertungsmethoden", "haftungsverhältnisse",
        "erläuterungen zur bilanz", "anlagevermögen", "jahresabschluss",
    ]) and "anhang" in text_lower:
        return "anhang"

    return "unknown"


def _detect_excel_type(path: Path) -> str:
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception:
        return "unknown"

    # Alle Zellen der ersten beiden Blätter durchsuchen (max 20 Zeilen)
    text = ""
    for ws in list(wb.worksheets)[:2]:
        for row in ws.iter_rows(max_row=20, values_only=True):
            text += " ".join(str(v) for v in row if v is not None) + "\n"
    wb.close()

    text_lower = text.lower()

    if any(kw in text_lower for kw in [
        "personalnummer", "abrechnungskreis", "köpfe", "kopfe",
        "mitarbeiter", "arbeitnehmer", "dienstnehmer", "personal",
    ]):
        return "hr_employees"

    return "unknown"


# ---------------------------------------------------------------------------
# Extraktion: Anhang-Positionen
# ---------------------------------------------------------------------------
def extract_from_anhang(pdf_path: Path) -> list[AnhangPosition]:
    """
    Extrahiert prüfbare Positionen aus dem Anhang:
      - Haftungsverhältnisse (Gesamtbetrag)
      - Arbeitnehmer (Gesamt, durchschnittlich)
    """
    positions: list[AnhangPosition] = []
    pdf_path = Path(pdf_path)

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page_num, page in enumerate(pdf.pages, 1):
            text = page.extract_text() or ""
            lines = text.split("\n")

            _extract_haftungen(lines, page_num, positions)
            _extract_arbeitnehmer(lines, page_num, positions)

    # Duplikate eliminieren (gleiche section + label, ersten Treffer behalten)
    seen: set[tuple[str, str]] = set()
    unique: list[AnhangPosition] = []
    for p in positions:
        key = (p.section, p.label)
        if key in seen:
            continue
        seen.add(key)
        unique.append(p)
    return unique


def _extract_haftungen(lines: list[str], page: int, out: list[AnhangPosition]) -> None:
    """Sucht 'Gesamtbetrag der Haftungsverhältnisse' und liest die EUR-Beträge."""
    for i, line in enumerate(lines):
        if re.search(r"Gesamtbetrag\s+der\s+Haftungsverh", line, re.I):
            nums = _trailing_numbers(line)
            # Manchmal stehen die Zahlen in der nächsten Zeile
            if not nums and i + 1 < len(lines):
                nums = _trailing_numbers(lines[i + 1])
            if nums:
                out.append(AnhangPosition(
                    section="Haftungsverhaeltnisse",
                    label="Gesamtbetrag der Haftungsverhältnisse",
                    current_value=nums[0],
                    prior_value=nums[1] if len(nums) > 1 else None,
                    page=page,
                    source_line=line.strip(),
                ))


def _extract_arbeitnehmer(lines: list[str], page: int, out: list[AnhangPosition]) -> None:
    """
    Sucht die Arbeitnehmer-Tabelle und extrahiert Gesamt-Zeile.
    Typisches Format:
        Arbeiter   6    7
        Angestellte 200  216
        Gesamt     206  223
    """
    # Kontext-Fenster: prüfen ob Arbeitnehmer/Angestellte-Sektion
    full_text = "\n".join(lines).lower()
    if not any(kw in full_text for kw in [
        "arbeitnehmer", "angestellte", "arbeiter", "dienstnehmer",
        "durchschnittlich", "mitarbeiter",
    ]):
        return

    for i, line in enumerate(lines):
        # Zeile die mit "Gesamt" beginnt und Zahlen hat
        if re.match(r"\s*Gesamt\b", line, re.I):
            nums = _trailing_numbers(line)
            if not nums:
                nums = _all_numbers(line)
            if nums and nums[0] > 0:
                # Sicherstellen, dass wir im richtigen Abschnitt sind
                ctx = "\n".join(lines[max(0, i - 15):i]).lower()
                if any(kw in ctx for kw in [
                    "arbeitnehmer", "angestellte", "arbeiter",
                    "dienstnehmer", "durchschnittlich",
                ]):
                    out.append(AnhangPosition(
                        section="Arbeitnehmer",
                        label="Durchschnittliche Arbeitnehmer (Gesamt)",
                        current_value=nums[0],
                        prior_value=nums[1] if len(nums) > 1 else None,
                        page=page,
                        source_line=line.strip(),
                        note=(
                            "Achtung: Anhang weist DURCHSCHNITTLICHE Zahl aus. "
                            "Detailunterlage kann STICHTAGSBESTAND sein – "
                            "methodischer Unterschied möglich."
                        ),
                    ))


# ---------------------------------------------------------------------------
# Extraktion: Bankgarantien / Haftungsobligo (PDF)
# ---------------------------------------------------------------------------
def extract_from_bank_guarantees(pdf_path: Path) -> list[ExtractedFact]:
    """
    Extrahiert die Gesamtsumme aus einer Bank-Garantieaufstellung.
    Erkennt Zeilen wie 'SUMME EURGGW 796.448,65'.
    """
    facts: list[ExtractedFact] = []
    pdf_path = Path(pdf_path)

    # Alle Summen-Kandidaten sammeln, dann den größten (letzten) nehmen
    candidates: list[tuple[float, str]] = []

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.split("\n"):
                line_up = line.upper()
                # WICHTIG: nur Bank-spezifische Summen, keine allgemeine "SUMME"
                # (sonst fängt dieser Extraktor "SUMME RÜCKSTELLUNGEN" etc. ab).
                if any(kw in line_up for kw in [
                    "SUMME EURGGW", "SUMME EUR ", "GESAMT-SUMME",
                    "GESAMTOBLIGO", "SUMME OBLIGO", "SUMME HAFTUNG",
                    "SUMME GARANTIE", "SUMME BANKGARANTIE",
                ]) or line_up.strip() in ("SUMME EUR", "TOTAL EUR"):
                    nums = _trailing_numbers(line)
                    if nums and nums[-1] > 0:
                        candidates.append((nums[-1], line.strip()))

    if candidates:
        # Nehme den größten Wert (Gesamtsumme, nicht Zwischensummen)
        best_val, best_line = max(candidates, key=lambda x: x[0])
        facts.append(ExtractedFact(
            source_type="bank_guarantees",
            label="Gesamtbetrag Bankgarantien / Haftungsobligo",
            value=best_val,
            filename=pdf_path.name,
            source_line=best_line,
        ))

    return facts


# ---------------------------------------------------------------------------
# Extraktion: Personalstand / Mitarbeiter (Excel)
# ---------------------------------------------------------------------------
def extract_from_hr_excel(xlsx_path: Path,
                          entity_codes: tuple = ()) -> list[ExtractedFact]:
    """
    Extrahiert den Mitarbeiterstand aus einer HR-Excel-Datei.

    Strategie:
    1. Pivot/Zusammenfassungs-Sheet bevorzugen (erkennt an Header mit 'Köpfe'/'Anzahl')
    2. Zeile der geprüften Gesellschaft suchen – erkennbar an einer der
       ``entity_codes`` (Kurzname oder Firmen-Code). Diese Kennungen sind
       mandantenspezifisch und kommen aus dem Mandanten-Plugin
       (``Pipeline.hr_entity_codes``), nicht aus diesem Modul.
    3. Die Köpfe/Anzahl-Spalte auslesen – KEIN Summing über alle Zeilen, nur
       die eine passende Zeile (oder das Gesamtergebnis, wenn keine Kennung
       hinterlegt ist oder keine Zeile darauf passt).

    Args:
        entity_codes: Kennungen der geprüften Gesellschaft, z.B. ``("XY", "100")``.
            Leer (Standard) -> es zählt allein die Gesamt-/Summenzeile.
    """
    facts: list[ExtractedFact] = []
    xlsx_path = Path(xlsx_path)

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    except Exception as e:
        return facts

    # Pivot-Sheet bevorzugen (kurze Zusammenfassung), Details-Sheet überspringen
    pivot_kws = {"pivot", "zusammenfassung", "summary", "übersicht", "uebersicht", "auswertung"}
    detail_kws = {"details", "detail", "liste", "rohdaten", "raw"}

    sheets_ordered = sorted(
        wb.worksheets,
        key=lambda ws: (
            0 if any(kw in ws.title.lower() for kw in pivot_kws) else
            2 if any(kw in ws.title.lower() for kw in detail_kws) else 1
        )
    )

    for ws in sheets_ordered:
        rows = list(ws.iter_rows(values_only=True))
        if not rows or len(rows) > 50:
            # Details-Sheets haben viele Zeilen → überspringen
            # (wir wollen nur Pivot/Zusammenfassung mit wenigen Zeilen)
            continue

        # Spaltenkopf finden
        header_row_idx = -1
        count_col_idx  = -1
        for ri, row in enumerate(rows[:5]):
            vals_lower = [str(v).lower() if v is not None else "" for v in row]
            if any("köpfe" in v or "kopfe" in v or "anzahl" in v
                   or "mitarbeiter" in v or "stand" in v for v in vals_lower):
                header_row_idx = ri
                # Zähler-Spalte: letzte Spalte im Header mit passendem Keyword
                for ci, val in enumerate(row):
                    if val is not None and any(
                        kw in str(val).lower()
                        for kw in ["köpfe", "kopfe", "anzahl", "mitarbeiter", "stand", "count"]
                    ):
                        count_col_idx = ci
                break

        # Zeilen nach der geprüften Gesellschaft oder der Summenzeile durchsuchen
        codes_upper = tuple(str(c).strip().upper() for c in entity_codes if str(c).strip())
        for ri, row in enumerate(rows):
            if ri == header_row_idx:
                continue
            vals_str = [str(v).strip() if v is not None else "" for v in row]
            full_upper = " ".join(vals_str).upper()

            is_entity = bool(codes_upper) and any(
                v.upper() in codes_upper
                for v in vals_str if v
            )
            is_total = any(
                kw in full_upper
                for kw in ["GESAMTERGEBNIS", "GESAMT", "TOTAL", "SUMME"]
            )

            if not (is_entity or is_total):
                continue

            # Wert aus der Köpfe-Spalte lesen
            count_val: Optional[float] = None
            if count_col_idx >= 0 and count_col_idx < len(row):
                v = row[count_col_idx]
                if isinstance(v, (int, float)) and v > 0:
                    count_val = float(v)

            # Fallback: letzter Integer-Wert in der Zeile
            if count_val is None:
                for v in reversed(row):
                    if isinstance(v, (int, float)) and v > 0 and v < 10000 and v == int(v):
                        count_val = float(v)
                        break

            if count_val is not None:
                label = ("Mitarbeiterstand geprüfte Gesellschaft" if is_entity
                         else "Mitarbeiterstand Gesamt")
                facts.append(ExtractedFact(
                    source_type="hr_employees",
                    label=label,
                    value=count_val,
                    filename=xlsx_path.name,
                    source_line=" | ".join(vals_str),
                    sheet=ws.title,
                ))

        if facts:
            break   # Pivot-Sheet gefunden und ausgewertet → fertig

    wb.close()

    # Deduplizieren: die Zeile der geprüften Gesellschaft schlägt die Summenzeile
    eigene = [f for f in facts if "geprüfte Gesellschaft" in f.label]
    return eigene[:1] if eigene else facts[:1]
