# -*- coding: utf-8 -*-
"""Hirn-Erweiterungen: AAB-Ende, 12-Spalten-Anlagenspiegel, Inline-Vorjahr, Text-Sortierung."""
from pathlib import Path

from anhangspruefer.vorjahresvergleich.extractor import (
    AnhangItem,
    _extract_inline_vorjahr,
    _is_anhang_end,
    anhang_page_range,
)
from anhangspruefer.vorjahresvergleich.text_compare import TextRow, sort_text_rows


def test_aab_endet_den_anhang():
    assert _is_anhang_end("Allgemeine Auftragsbedingungen")
    assert _is_anhang_end("Allgemeine Auftragsbedingungen fuer Wirtschaftstreuhandberufe")
    pages = ["Deckblatt", "Anhang", "Bewertung", "Allgemeine Auftragsbedingungen fuer WT"]
    assert anhang_page_range(pages) == (1, 3)


def test_inline_vorjahr_paar():
    text = "Im Geschaeftsjahr waren im Durchschnitt 4 Arbeitnehmer (Vorjahr: 3 Arbeitnehmer) beschaeftigt."
    items = _extract_inline_vorjahr(text, 2)
    assert len(items) == 1
    assert items[0].current_values == [4.0]
    assert items[0].prior_values == [3.0]


def test_text_rows_aenderungen_zuerst():
    rows = [
        TextRow("a", "a", "IDENT", 1, 1),
        TextRow("neu", "", "NEU", 2, None),
        TextRow("x", "y", "GEÄNDERT", 3, 3),
        TextRow("", "alt", "FEHLT", None, 4),
    ]
    ordered = [r.status for r in sort_text_rows(rows)]
    assert ordered == ["FEHLT", "GEÄNDERT", "NEU", "IDENT"]


def test_excel_hat_getrennte_textblaetter(tmp_path):
    from openpyxl import load_workbook
    from anhangspruefer.vorjahresvergleich.comparator import CompareResult
    from anhangspruefer.vorjahresvergleich.excel_report import generate_excel

    result = CompareResult(
        current_pdf=Path("aktuell.pdf"),
        prior_pdf=Path("vorjahr.pdf"),
        rows=[],
        text_rows=[
            TextRow("neu hier", "", "NEU", 2, None),
            TextRow("", "alt dort", "FEHLT", None, 3),
            TextRow("jetzt so", "vorher so", "GEÄNDERT", 4, 4),
            TextRow("gleich", "gleich", "IDENT", 1, 1),
        ],
    )
    out = tmp_path / "v.xlsx"
    generate_excel(result, out)
    wb = load_workbook(out)
    names = wb.sheetnames
    assert "Neu im Bericht" in names
    assert "Fehlt gegenüber Vorjahr" in names
    assert "Geänderter Text" in names
    assert "Textvergleich" not in names
    assert "Textänderungen" not in names
    assert wb["Neu im Bericht"]["B3"].value == "neu hier"
    assert wb["Fehlt gegenüber Vorjahr"]["B3"].value == "alt dort"
    assert wb["Geänderter Text"]["E3"].value == "jetzt so"
    assert wb["Übersicht"]["B15"].value == 1
    assert wb["Übersicht"]["B16"].value == 1
    assert wb["Übersicht"]["B17"].value == 1


def test_ja_page_range_bis_vor_aab():
    from anhangspruefer.vorjahresvergleich.extractor import aab_page_index, ja_page_range
    aab = "Allgemeine Auftragsbedingungen fuer Wirtschaftstreuhandberufe\n" + ("Klausel " * 600)
    pages = ["Deckblatt", "Bilanz AKTIVA", "Anhang Angaben", aab]
    assert aab_page_index(pages) == 3
    assert ja_page_range(pages) == (0, 3)


def test_aab_index_ignores_inhaltsverzeichnis():
    from anhangspruefer.vorjahresvergleich.extractor import aab_page_index
    aab = "Allgemeine Auftragsbedingungen fuer Wirtschaftstreuhandberufe\n" + ("Klausel " * 600)
    pages = [
        "Inhaltsverzeichnis\nBilanz 4\nAllgemeine Auftragsbedingungen 13",
        "Bilanz AKTIVA Software",
        aab,
    ]
    assert aab_page_index(pages) == 2


def test_bilanz_zwei_zahlen_ohne_spaltenkopf(tmp_path, monkeypatch):
    """AKTIVA-Seite ohne EUR-EUR-Kopf: zwei Trailing-Zahlen = GJ | VJ."""
    from anhangspruefer.vorjahresvergleich import extractor as extr
    seite = "AKTIVA\nKassenbestand 12.341,43 13.341,43\n"
    monkeypatch.setattr(extr, "load_page_texts", lambda p, **k: [seite])
    items = extr.extract_items(tmp_path / "b.pdf", page_range=(0, None))
    kasse = [it for it in items if "Kassenbestand" in it.label]
    assert kasse, [it.label for it in items]
    assert kasse[0].current_values == [12341.43]
    assert kasse[0].prior_values == [13341.43]


def test_excel_hat_zahlen_nur_blaetter(tmp_path):
    from openpyxl import load_workbook
    from anhangspruefer.vorjahresvergleich.comparator import CompareResult, ComparisonRow
    from anhangspruefer.vorjahresvergleich.excel_report import generate_excel

    result = CompareResult(
        current_pdf=Path("aktuell.pdf"),
        prior_pdf=Path("vorjahr.pdf"),
        rows=[
            ComparisonRow("NeuPosten", 1, 10.0, None, 1, None, 0.0, "NUR_AKTUELL"),
            ComparisonRow("AltPosten", 1, None, 20.0, 0, 2, 0.0, "NUR_VORJAHR"),
            ComparisonRow("Passt", 1, 5.0, 5.0, 1, 1, 1.0, "OK"),
        ],
    )
    out = tmp_path / "z.xlsx"
    generate_excel(result, out)
    wb = load_workbook(out)
    assert "Nur aktuell" in wb.sheetnames
    assert "Nur Vorjahr" in wb.sheetnames
    assert wb["Nur aktuell"]["A2"].value == "NeuPosten"
    assert wb["Nur Vorjahr"]["A2"].value == "AltPosten"
