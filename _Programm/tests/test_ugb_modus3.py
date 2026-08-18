"""Tests für Modus 3: Loader-Blattwahl + Relevanz-Filter (bedingte Angaben).

Nur synthetische Daten – keine Mandanten-/KPMG-Unterlagen.
"""

from datetime import datetime

import openpyxl

from anhangspruefer.models.checklist import Checklist, ChecklistItem
from anhangspruefer.models.finding import Finding, ReviewResult, EvidenceItem
from anhangspruefer.models.enums import ComplianceStatus
from anhangspruefer.compliance.knowledge.checklist_loader import ChecklistLoader
from anhangspruefer.compliance.knowledge.relevance import (
    _detect_legal_form,
    _item_applicable,
    _size_class_applicable,
    apply_relevance,
    category_applicable,
)


# --- Loader: das richtige Blatt wählen (nicht das "Start"-Übersichtsblatt) ---
def test_load_from_xlsx_picks_master_sheet(tmp_path):
    wb = openpyxl.Workbook()
    start = wb.active
    start.title = "Start"
    start.append(["#", "Kategorie", "Prüfpunkte", "relevant?"])   # KEIN ID/Prüffrage
    start.append(["1", "Allgemein", "7", "Ja"])
    master = wb.create_sheet("UGB-Prüfprogramm")
    master.append(["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter", "Pflicht", "Anwendbar auf"])
    master.append(["K001", "Vorräte", "Angabe des Unterschiedsbetrags", "§ 238", "Vorräte", "Ja", "alle"])
    master.append(["K002", "Forderungen", "Erläuterung der Forderungen", "§ 225", "Forderung", "Ja", "alle"])
    p = tmp_path / "pp.xlsx"
    wb.save(p)

    cl = ChecklistLoader().load_from_xlsx(p)
    assert len(cl.items) == 2                       # nicht 0 (Bug: las 'Start')
    assert cl.get_item("K001").category == "Vorräte"


# --- Relevanz: allgemeine Kategorien immer, positionsabhängige nur bei Vorkommen ---
def test_category_applicable_general_is_always_true():
    assert category_applicable("Allgemein", "") is True
    assert category_applicable("Allgemeine Angaben", "beliebig") is True


def test_category_applicable_conditional():
    assert category_applicable("Vorräte", "Die Vorräte werden bewertet.".lower()) is True
    assert category_applicable("Vorräte", "hier steht nichts dazu") is False
    assert category_applicable("Anteilsbasierte Vergütungen", "keine") is False
    assert category_applicable("Umgründung Allgemein", "im Zuge der Verschmelzung".lower()) is True


def _finding(iid):
    return Finding(checklist_item_id=iid, status=ComplianceStatus.PARTIALLY_COMPLIANT, ugb_references=[])


def test_apply_relevance_marks_absent_positions_not_applicable():
    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(item_id="A1", category="Vorräte", description="x"))
    cl.add_item(ChecklistItem(item_id="B1", category="Anteilsbasierte Vergütungen", description="y"))
    cl.add_item(ChecklistItem(item_id="C1", category="Allgemein", description="z"))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    for iid in ("A1", "B1", "C1"):
        res.add_finding(_finding(iid))

    summary = apply_relevance(res, cl, "Die Vorräte werden zu Anschaffungskosten bewertet.")
    st = {f.checklist_item_id: f.status for f in res.findings}

    assert st["A1"] == ComplianceStatus.PARTIALLY_COMPLIANT   # Vorräte vorhanden -> bleibt
    assert st["B1"] == ComplianceStatus.NOT_APPLICABLE        # anteilsbasiert fehlt -> N/A
    assert st["C1"] == ComplianceStatus.PARTIALLY_COMPLIANT   # allgemein immer relevant
    assert "Anteilsbasierte Vergütungen" in summary["nicht_anwendbar"]
    assert summary["umgestellt"] == 1


# --- Rechtsform & Größenklasse (KPMG-Spalte) ---------------------------------
def test_detect_legal_form():
    assert _detect_legal_form("die musterfirma handels gmbh mit stammkapital") == "gmbh"
    assert _detect_legal_form("die muster aktiengesellschaft mit grundkapital") == "ag"
    assert _detect_legal_form("unklarer text ohne rechtsform") is None


def test_size_class_excludes_ag_only_items_for_gmbh():
    ag_only = ChecklistItem(item_id="K044", category="Allgemein", description="Aktiengattungen",
                            size_classes=["AG groß", "AG mittel"])
    mixed = ChecklistItem(item_id="K048", category="Allgemein", description="x",
                          size_classes=["AG groß", "GmbH groß"])
    unrestricted = ChecklistItem(item_id="K001", category="Allgemein", description="y")
    assert _size_class_applicable(ag_only, "gmbh", "gross") is False
    assert _size_class_applicable(mixed, "gmbh", "gross") is True
    assert _size_class_applicable(unrestricted, "gmbh", "gross") is True
    assert _size_class_applicable(ag_only, None, "gross") is True      # Rechtsform unklar -> nicht filtern
    assert _size_class_applicable(ag_only, "gmbh", None) is True       # Größenklasse unklar -> nicht filtern


def test_size_class_parser_all_combinations():
    """Parser deckt alle (Rechtsform, Größe)-Kombinationen ab; unbekannte
    Größe der eigenen Rechtsform schließt die Anwendbarkeit klar aus."""
    item = ChecklistItem(item_id="K900", category="Allgemein", description="x",
                         size_classes=["AG groß; AG mittel; GmbH groß; GmbH mittel"])
    for form in ("gmbh", "ag"):
        for size in ("groß", "mittel"):
            size_key = "gross" if size == "groß" else size
            assert _size_class_applicable(item, form, size_key) is True
    assert _size_class_applicable(item, "gmbh", "klein") is False
    assert _size_class_applicable(item, "ag", "klein") is False
    # unparsbare Größenklasse -> konservativ nicht filtern
    unparsbar = ChecklistItem(item_id="K901", category="Allgemein", description="y",
                              size_classes=["Sonderfall siehe Hinweis"])
    assert _size_class_applicable(unparsbar, "gmbh", "klein") is True


def test_item_trigger_verbrauchsfolge_hat_vorrang_vor_wertpapier():
    # K022-Fall: Frage nennt Verbrauchsfolgeverfahren UND (beiläufig) Wertpapiere.
    item = ChecklistItem(
        item_id="K022", category="Allgemein",
        description=("Angabe des Unterschiedsbetrages zu Börsenkursen bei Anwendung "
                     "eines Verbrauchsfolgeverfahrens gem. § 209 (2) UGB, auch bei Wertpapieren"))
    # Mandant OHNE Verbrauchsfolgeverfahren -> nicht anwendbar
    assert _item_applicable(item, "die waren werden zu anschaffungskosten bewertet") is False
    # Mandant MIT FIFO -> anwendbar (auch wenn 'wertpapier' im Abschluss fehlt)
    assert _item_applicable(item, "die vorräte werden nach dem fifo-verfahren bewertet") is True


def test_item_trigger_hybridinstrumente():
    item = ChecklistItem(item_id="K060", category="Allgemein",
                         description="Angabe des Bestehens von Genussscheinen, Wandelschuldverschreibungen oder vergleichbaren Wertpapieren")
    assert _item_applicable(item, "text ohne solche instrumente") is False
    assert _item_applicable(item, "es bestehen genussrechte aus dem jahr 2020") is True


def test_item_trigger_festverzinsliche_wertpapiere():
    item = ChecklistItem(item_id="K014", category="Allgemein",
                         description="Beschreibung der Methoden bei festverzinslichen Wertpapieren")
    assert _item_applicable(item, "kein einschlägiger text") is False
    assert _item_applicable(item, "die wertpapiere des anlagevermögens ...") is True
    # 'derivativer Firmenwert' ist KEIN Derivat -> der Derivate-Trigger darf
    # nicht greifen; maßgeblich ist, ob ein FIRMENWERT bilanziert ist.
    fw = ChecklistItem(item_id="K018", category="Allgemein",
                       description="Zuordnung eines derivativen Geschäfts-(Firmen-)werts")
    assert _item_applicable(fw, "der firmenwert wird linear abgeschrieben") is True
    # Ohne Firmenwert im Abschluss ist die Angabe nicht anwendbar (Prüfervorgabe)
    assert _item_applicable(fw, "text ohne solche position") is False


def test_loader_reads_size_classes(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UGB-Prüfprogramm"
    ws.append(["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter", "Pflicht",
               "Anwendbar auf", "Größenklasse (KPMG)"])
    ws.append(["K044", "Allgemein", "Aktienangaben", "§ 241", "Aktien", "Ja",
               "alle", "AG groß; AG mittel"])
    p = tmp_path / "pp.xlsx"
    wb.save(p)
    cl = ChecklistLoader().load_from_xlsx(p)
    assert cl.get_item("K044").size_classes == ["AG groß", "AG mittel"]
    assert cl.get_item("K044").applicable_to == ["alle"]


# --- Ausgefüllte KPMG-Checkliste als Excel ---
def test_generate_checklist_xlsx(tmp_path):
    import openpyxl
    from anhangspruefer.compliance.reporting.checklist_excel import generate_checklist_xlsx

    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(item_id="K1", category="Vorräte", description="Angabe X", ugb_references=["§ 238"]))
    cl.add_item(ChecklistItem(item_id="K2", category="Anteilsbasierte Vergütungen", description="Angabe Y"))
    cl.add_item(ChecklistItem(item_id="K3", category="Verbindlichkeiten",
                              description="Angabe der Restlaufzeit über 5 Jahre"))
    res = ReviewResult(document_name="Anhang.pdf", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(Finding(checklist_item_id="K1", status=ComplianceStatus.COMPLIANT, ugb_references=["§ 238"]))
    f2 = Finding(checklist_item_id="K2", status=ComplianceStatus.NOT_APPLICABLE, ugb_references=[])
    f2.add_evidence(EvidenceItem(section_id="s", section_title="Anhang", quote="schwacher treffer", page_number=5))
    res.add_finding(f2)
    f3 = Finding(checklist_item_id="K3", status=ComplianceStatus.NOT_COMPLIANT, ugb_references=[])
    f3.missing_elements = ["Gesamtbetrag Restlaufzeit über 5 Jahre nicht angegeben"]
    res.add_finding(f3)

    out = tmp_path / "cl.xlsx"
    generate_checklist_xlsx(cl, res, out)

    wb = openpyxl.load_workbook(out)
    assert "KPMG-Checkliste (ausgefüllt)" in wb.sheetnames
    ws = wb["KPMG-Checkliste (ausgefüllt)"]
    assert ws.max_row == 4                          # Kopf + 3 Punkte
    row2 = [c.value for c in ws[2]]
    row3 = [c.value for c in ws[3]]
    row4 = [c.value for c in ws[4]]
    assert row2[1] == "K1" and row2[5] == "Ja"      # klares Verdikt, kein "Teilweise"
    assert row3[1] == "K2" and row3[5] == "n. a. (Rechtsgrund)"
    assert (row3[6] or "") == ""                    # bei n. a. kein Nachweis
    # Bei "Fehlt": klare, KONKRETE Aussage was fehlt
    assert row4[5] == "Fehlt"
    assert row4[7] == "Fehlt: Gesamtbetrag Restlaufzeit über 5 Jahre nicht angegeben"


def test_checklist_xlsx_has_pruefer_override(tmp_path):
    """Prüfer kann übersteuern: Dropdown 'Prüfer-Urteil', 'Final' = Prüfer schlägt Tool."""
    import openpyxl
    from anhangspruefer.compliance.reporting.checklist_excel import generate_checklist_xlsx

    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(item_id="K1", category="Vorräte", description="Angabe X"))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(Finding(checklist_item_id="K1",
                            status=ComplianceStatus.NOT_ASSESSABLE, ugb_references=[]))  # Tool: Offen
    out = tmp_path / "cl.xlsx"
    generate_checklist_xlsx(cl, res, out)

    wb = openpyxl.load_workbook(out)                 # ohne data_only -> Formeln sichtbar
    ws = wb["KPMG-Checkliste (ausgefüllt)"]
    header = [c.value for c in ws[1]]
    assert header[9] == "Prüfer-Urteil" and header[10] == "Final" and header[11] == "Prüfer-Kommentar"
    # Final-Formel: Prüfer-Urteil (J) übersteuert Tool-Ergebnis (interner Status, Spalte M –
    # trägt den plain-text Status unabhängig vom ausführlicheren Anzeigetext in F)
    assert ws.cell(row=2, column=11).value == '=IF(J2<>"",J2,M2)'
    # Dropdown mit den vier Verdikten vorhanden
    dvs = list(ws.data_validations.dataValidation)
    assert any(d.formula1 == '"Ja,Fehlt,n. a.,Offen"' for d in dvs)


# --- Loader: Kategorie vom Kategorieblatt, Prüfer-Vorgabe aus Blatt "Start" --
def _build_pp_with_category_sheets(tmp_path, start_relevant="Ja"):
    """Synthetisches Prüfprogramm mit Master- + Kategorieblättern + 'Start'."""
    wb = openpyxl.Workbook()
    start = wb.active
    start.title = "Start"
    start.append(["#", "Kategorie (Klick zum Blatt)", "Prüfpunkte", "relevant?"])
    start.append([1, "1 Allgemeine Angaben", 1, start_relevant])
    start.append([2, "2 Vorräte", 1, "Ja"])

    master = wb.create_sheet("UGB-Prüfprogramm")
    master.append(["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter", "Pflicht",
                   "Anwendbar auf", "Größenklasse (KPMG)"])
    master.append(["K001", "Allgemein", "Allgemeine Angabe", "§ 236", "Angabe", "Ja", "alle", ""])
    master.append(["K002", "Allgemein", "Angabe zu Vorräten", "§ 238", "Vorräte", "Ja", "alle", ""])

    kat1 = wb.create_sheet("1 Allgemeine Angaben")
    kat1.append(["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter", "Pflicht", "Anwendbar auf"])
    kat1.append(["K001", "Allgemein", "Allgemeine Angabe", "§ 236", "Angabe", "Ja", "alle"])

    kat2 = wb.create_sheet("2 Vorräte")
    kat2.append(["ID", "Kategorie", "Prüffrage", "UGB-§", "Stichwörter", "Pflicht", "Anwendbar auf"])
    kat2.append(["K002", "Allgemein", "Angabe zu Vorräten", "§ 238", "Vorräte", "Ja", "alle"])

    p = tmp_path / "pp.xlsx"
    wb.save(p)
    return p


def test_loader_derives_category_from_sheet_not_master_column(tmp_path):
    p = _build_pp_with_category_sheets(tmp_path)
    cl = ChecklistLoader().load_from_xlsx(p)
    # Master-Spalte "Kategorie" ist bei beiden Punkten die Sammelkategorie
    # "Allgemein" -- die WAHRE Kategorie kommt vom Blattnamen (ohne Nummer).
    assert cl.get_item("K001").category == "Allgemeine Angaben"
    assert cl.get_item("K002").category == "Vorräte"


def test_loader_reads_start_sheet_relevance(tmp_path):
    p = _build_pp_with_category_sheets(tmp_path, start_relevant="Nein")
    cl = ChecklistLoader().load_from_xlsx(p)
    assert cl.get_item("K001").pruefer_relevant is False   # Blatt "relevant?" = Nein
    assert cl.get_item("K002").pruefer_relevant is True    # andere Kategorie unberührt


# --- Falsche Entlastung: Größenklassen-Pflicht darf nicht durch Stichwort- --
# --- Abwesenheit überschrieben werden; maschinelles n.a. trägt Präfix. -----
def test_size_class_mandatory_point_never_gets_rechtsgrund_na():
    """Ein für GmbH klein PFLICHT-Punkt darf NIE per Rechtsgrund (Größen-
    klasse) entlastet werden – auch bei magerem Anhangtext. Greift zusätzlich
    die (separate) Stichwort-Heuristik, muss der Grund als 'Maschinell n. a.'
    gekennzeichnet sein, NIE als scheinbar rechtssicheres 'Nicht erforderlich
    für ...' (das wäre die falsche Entlastung aus der Baseline-Messung)."""
    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(
        item_id="K900", category="Allgemein",
        description="Angabe und Begründung, wenn die Darstellungsstetigkeit nicht beibehalten wird",
        size_classes=["AG groß; AG mittel; AG klein; GmbH groß; GmbH mittel; GmbH klein"],
    ))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(_finding("K900"))

    apply_relevance(res, cl, "ein sehr magerer anhangtext ohne besondere begriffe",
                    legal_form="gmbh", size_class="klein")
    f = res.findings[0]
    reason = f.technical_reasoning or ""
    assert "Nicht erforderlich für" not in reason        # nie fälschlich per Rechtsgrund entlastet
    if f.status == ComplianceStatus.NOT_APPLICABLE:
        assert reason.startswith("Maschinell n. a. – bitte stichprobenweise prüfen:")
    else:
        assert f.status == ComplianceStatus.PARTIALLY_COMPLIANT   # unberührt


def test_size_class_rechtsgrund_na_for_excluded_size():
    """Ein Punkt, der laut KPMG-Größenklasse NICHT für GmbH klein gilt, wird
    per Rechtsgrund (nicht maschinell) auf n.a. gesetzt."""
    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(
        item_id="K905", category="Allgemein", description="Angabe X",
        size_classes=["AG groß; AG mittel; GmbH groß; GmbH mittel"],   # ohne 'klein'
    ))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(_finding("K905"))

    apply_relevance(res, cl, "beliebiger text", legal_form="gmbh", size_class="klein")
    f = res.findings[0]
    assert f.status == ComplianceStatus.NOT_APPLICABLE
    assert not f.technical_reasoning.startswith("Maschinell n. a.")
    assert "GmbH klein" in f.technical_reasoning


def test_machine_na_reason_carries_prefix():
    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(item_id="B1", category="Anteilsbasierte Vergütungen", description="y"))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(_finding("B1"))

    apply_relevance(res, cl, "text ohne solche instrumente")
    f = res.findings[0]
    assert f.status == ComplianceStatus.NOT_APPLICABLE
    assert f.technical_reasoning.startswith("Maschinell n. a. – bitte stichprobenweise prüfen:")


def test_rechtsgrund_na_reason_has_no_machine_prefix():
    cl = Checklist(name="t", version="")
    cl.add_item(ChecklistItem(item_id="K001", category="Allgemein", description="Ausschließlich Start-relevant",
                              pruefer_relevant=False))
    res = ReviewResult(document_name="d", checklist_name="t", review_timestamp=datetime(2026, 1, 1))
    res.add_finding(_finding("K001"))

    apply_relevance(res, cl, "beliebiger text")
    f = res.findings[0]
    assert f.status == ComplianceStatus.NOT_APPLICABLE
    assert not f.technical_reasoning.startswith("Maschinell n. a.")
